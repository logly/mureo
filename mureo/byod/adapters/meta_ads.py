"""Meta Ads BYOD adapter — consumes the user's Ads Manager Excel export.

Meta does not offer a "scripts inside the Ads UI" runtime equivalent
to Google Ads Scripts, and bundling a mureo-managed Marketing API
client would violate the project's "no SaaS" contract. The pragmatic
no-OAuth path is the manual XLSX export users can run themselves
from Ads Manager:

  Ads Manager → Reports → Customize → set columns → Export → Excel

This adapter normalizes that single-tab export into the 4 CSVs the
existing ``ByodMetaAdsClient`` (``mureo/byod/clients.py``) reads
under ``~/.mureo/byod/meta_ads/``:

  campaigns.csv      campaign_id, name, status, objective, daily_budget_jpy
  ad_sets.csv        ad_set_id, campaign_id, name, status
  ads.csv            ad_id, ad_set_id, name, status
  metrics_daily.csv  date, campaign_id, impressions, clicks, cost_jpy, conversions

Identity is synthesized from name (deterministic SHA-256 hash) because
the Ads Manager export does not include numeric IDs by default. The
hash is stable across re-imports so STATE.json references continue
to resolve.

Recognized header names — multilingual (English / 日本語 / Español /
Português / 한국어 / 繁體中文 / 简体中文 / Français / Deutsch). The
adapter looks for any of the per-language aliases on each required
column, so a workbook exported with Ads Manager in Japanese
(キャンペーン名, インプレッション, …) imports the same way as an English
export.

Currency: only JPY is accepted. A non-JPY symbol prefix in the spend
column (``$``, ``€``, ``£``, ``₩``, ``₹``, ``¢``) raises
``UnsupportedFormatError`` so the user fixes Account currency before
the BYOD pipeline silently mis-reports cost.

Pivot tables: Reports section pivot exports include subtotal rows
where the date column reads ``All`` (or is blank) — these are
filtered out by ``_parse_day`` returning the empty string, so the
metric loop ``if not day: continue`` simply skips them. Detail rows
(date = ``YYYY-MM-DD``) are kept and aggregated per ``(day, campaign)``.
"""

from __future__ import annotations

import csv as _csv
import hashlib
import re
from dataclasses import dataclass
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.workbook.workbook import Workbook

SOURCE_FORMAT = "mureo_meta_ads_export_v1"


@dataclass
class ImportResult:
    """Per-platform return value from a normalize_from_workbook call."""

    rows: int
    date_range: tuple[str, str]
    files_written: list[str]
    source_format: str
    campaigns: int
    ad_groups: int  # repurposed as ad_sets count for Meta


class UnsupportedFormatError(ValueError):
    """Raised when no sheet in the workbook matches the expected schema."""


# ---------------------------------------------------------------------------
# Header alias map — lowercased, stripped, multiple Ads Manager wordings →
# canonical column names used internally by this adapter.
# ---------------------------------------------------------------------------

# All aliases lowercased + stripped to match what `_header_row` produces.
# Verified against actual Meta Ads Manager exports in 9 locales (`en`,
# `ja`, `zh_CN`, `zh_TW`, `ko_KR`, `es_ES`, `pt_BR`, `de_DE`, `fr_FR`).
# When Meta updates UI translations, update each tuple from a fresh
# locale export rather than guessing.
_DATE_ALIASES = (
    # English (verified)
    "day",
    "reporting starts",
    "date",
    # Japanese (verified)
    "日",
    "レポート開始日",
    # Simplified Chinese (verified)
    "报告开始日期",
    # Traditional Chinese (verified)
    "分析報告開始",
    # Korean (verified)
    "보고 시작",
    # Spanish (verified)
    "inicio del informe",
    # Portuguese BR (verified)
    "início dos relatórios",
    # German (verified)
    "berichtsstart",
    # French (verified)
    "début des rapports",
)

_CAMPAIGN_ALIASES = (
    # English (verified)
    "campaign name",
    # Japanese (verified)
    "キャンペーン名",
    # Simplified Chinese (verified)
    "广告系列名称",
    # Traditional Chinese (verified)
    "行銷活動名稱",
    # Korean (verified)
    "캠페인 이름",
    # Spanish (verified)
    "nombre de la campaña",
    # Portuguese BR (verified)
    "nome da campanha",
    # German (verified)
    "name der kampagne",
    # French (verified)
    "nom de la campagne",
)

# Optional column — verified from actual exports when present, plus
# best-effort translations for locales whose verification export
# didn't include this column. Unverified entries fall back to
# "no match" silently, which means ad-set names won't be extracted
# but the import still succeeds with campaign-level data.
_AD_SET_ALIASES = (
    "ad set name",  # English (verified)
    "広告セット名",  # Japanese (verified)
    "广告组名称",  # Simplified Chinese (best-effort)
    "廣告組合名稱",  # Traditional Chinese (best-effort)
    "광고 세트 이름",  # Korean (best-effort)
    "nombre del conjunto de anuncios",  # Spanish (best-effort)
    "nome do conjunto de anúncios",  # Portuguese BR (best-effort)
    "anzeigengruppenname",  # German (best-effort)
    "nom de l'ensemble de publicités",  # French (best-effort)
)

_AD_ALIASES = (
    "ad name",  # English (verified)
    "広告名",  # Japanese (verified)
    "广告名称",  # Simplified Chinese (best-effort)
    "廣告名稱",  # Traditional Chinese (best-effort)
    "광고 이름",  # Korean (best-effort)
    "nombre del anuncio",  # Spanish (best-effort)
    "nome do anúncio",  # Portuguese BR (best-effort)
    "anzeigenname",  # German (best-effort)
    "nom de la publicité",  # French (best-effort)
)

_IMPRESSIONS_ALIASES = (
    "impressions",  # English / French (verified)
    "インプレッション",  # Japanese (verified)
    "展示次数",  # Simplified Chinese (verified)
    "曝光次數",  # Traditional Chinese (verified)
    "노출",  # Korean (verified)
    "impresiones",  # Spanish (verified)
    "impressões",  # Portuguese BR (verified)
    "impressionen",  # German (verified)
)

# Optional column — same best-effort handling as ad set / ad name.
_CLICKS_ALIASES = (
    "clicks (all)",
    "link clicks",
    "clicks",
    # Japanese — verified actual export uses "クリック(すべて)" with
    # half-width parens and no space; older / customized exports have
    # also been observed using 全件 wording and full-width parens, so
    # both forms are aliased.
    "クリック(すべて)",
    "クリック (すべて)",
    "クリック(全件)",
    "クリック (全件)",
    "クリック数（全件）",
    "クリック数(全件)",
    "リンクのクリック",
    "リンククリック",
    "クリック数",
    # Other locales — best-effort, not yet verified from real exports.
    "全部点击",
    "链接点击量",
    "点击量",
    "全部點擊次數",
    "連結點擊次數",
    "點擊次數",
    "전체 클릭",
    "링크 클릭",
    "클릭",
    "clic (todos)",
    "clic en el enlace",
    "clics",
    "cliques (todos)",
    "cliques no link",
    "cliques",
    "klicks (alle)",
    "link-klicks",
    "klicks",
    "clics (tous)",
    "clics sur le lien",
)

_SPEND_ALIASES = (
    # English (verified) + non-JPY suffix variants for safety
    "amount spent (jpy)",
    "amount spent",
    "spend",
    # Japanese (verified)
    "消化金額 (jpy)",
    "消化金額（jpy）",
    "消化金額",
    # Simplified Chinese (verified)
    "已花费金额 (jpy)",
    "已花费金额",
    # Traditional Chinese (verified)
    "花費金額 (jpy)",
    "花費金額",
    # Korean (verified)
    "지출 금액 (jpy)",
    "지출 금액",
    # Spanish (verified)
    "importe gastado (jpy)",
    "importe gastado",
    # Portuguese BR (verified)
    "valor usado (jpy)",
    "valor usado",
    # German (verified)
    "ausgegebener betrag (jpy)",
    "ausgegebener betrag",
    # French (verified)
    "montant dépensé (jpy)",
    "montant dépensé",
)

_CONVERSIONS_ALIASES = (
    "results",  # English (verified)
    "conversions",
    "結果",  # Japanese (verified)
    "コンバージョン",
    "成效",  # Simplified Chinese (verified)
    "成果",  # Traditional Chinese (verified)
    "결과",  # Korean (verified)
    "전환",
    "resultados",  # Spanish / Portuguese BR (verified)
    "conversiones",
    "conversões",
    "ergebnisse",  # German (verified)
    "résultats",  # French (verified)
)

# Phase 3 — additional analytical columns. Verified in the same 9-locale
# export sweep that established the core column aliases.
_REACH_ALIASES = (
    "reach",  # English (verified)
    "リーチ",  # Japanese (verified)
    "覆盖人数",  # Simplified Chinese (verified)
    "觸及人數",  # Traditional Chinese (verified)
    "도달",  # Korean (verified)
    "alcance",  # Spanish / Portuguese BR (verified)
    "couverture",  # French (verified)
    "reichweite",  # German (verified)
)

_FREQUENCY_ALIASES = (
    "frequency",  # English (verified)
    "フリークエンシー",  # Japanese (verified)
    "频次",  # Simplified Chinese (best-effort)
    "頻率",  # Traditional Chinese (best-effort)
    "빈도",  # Korean (best-effort)
    "frecuencia",  # Spanish (best-effort)
    "frequência",  # Portuguese BR (best-effort)
    "fréquence",  # French (best-effort)
    "frequenz",  # German (best-effort)
)

_RESULT_INDICATOR_ALIASES = (
    "result indicator",  # English (verified)
    "結果インジケーター",  # Japanese (verified)
    "成效指标",  # Simplified Chinese (verified)
    "成果指標",  # Traditional Chinese (verified)
    "결과 표시 도구",  # Korean (verified)
    "indicador de resultado",  # Spanish (verified)
    "indicador de resultados",  # Portuguese BR (verified)
    "indikator für ergebnisse",  # German (verified)
    "indicateur de résultats",  # French (verified)
)

# Demographics breakdown columns — populated only when the user enables
# breakdowns in the Reports section. Skipped when the column is absent
# or every row has the locale's "All" sentinel value.
_AGE_ALIASES = (
    "age",  # English (best-effort)
    "年齢",  # Japanese (best-effort)
    "年龄",  # Simplified Chinese (best-effort)
    "年齡",  # Traditional Chinese (best-effort)
    "연령",  # Korean (best-effort)
    "edad",  # Spanish (best-effort)
    "idade",  # Portuguese BR (best-effort)
    "âge",  # French (best-effort)
    "alter",  # German (best-effort)
)

_GENDER_ALIASES = (
    "gender",  # English
    "性別",  # Japanese / Traditional Chinese
    "性别",  # Simplified Chinese
    "성별",  # Korean
    "género",  # Spanish
    "gênero",  # Portuguese BR
    "sexe",  # French
    "geschlecht",  # German
)

_REGION_ALIASES = (
    "region",  # English
    "地域",  # Japanese
    "地区",  # Simplified Chinese
    "區域",  # Traditional Chinese
    "지역",  # Korean
    "región",  # Spanish
    "região",  # Portuguese BR
    "région",  # French
)

_PLACEMENT_ALIASES = (
    "placement",  # English
    "配置",  # Japanese (verified) / Simplified Chinese
    "版位",  # Traditional Chinese
    "노출 위치",  # Korean
    "ubicación",  # Spanish
    "posicionamento",  # Portuguese BR
    "platzierung",  # German
    "placement publicitaire",  # French (best-effort)
)

_PLATFORM_ALIASES = (
    "platform",  # English
    "プラットフォーム",  # Japanese (verified)
    "平台",  # Simplified / Traditional Chinese (best-effort)
    "플랫폼",  # Korean (best-effort)
    "plataforma",  # Spanish / Portuguese BR (best-effort)
    "plateforme",  # French (best-effort)
    "plattform",  # German (best-effort)
)

_DEVICE_ALIASES = (
    "device platform",  # English
    "デバイスプラットフォーム",  # Japanese (verified)
    "device",
    "デバイス",
    "设备平台",  # Simplified Chinese (best-effort)
    "裝置平台",  # Traditional Chinese (best-effort)
    "기기 플랫폼",  # Korean (best-effort)
    "plataforma del dispositivo",  # Spanish (best-effort)
    "plataforma do dispositivo",  # Portuguese BR (best-effort)
    "plateforme de l'appareil",  # French (best-effort)
    "geräteplattform",  # German (best-effort)
)

# Creative-info columns — best-effort, present only in custom reports.
_CREATIVE_IMAGE_URL_ALIASES = (
    "image url",
    "image_url",
    "画像url",
    "图片网址",
)
_CREATIVE_VIDEO_URL_ALIASES = (
    "video url",
    "video_url",
    "動画url",
    "视频网址",
)
_CREATIVE_HEADLINE_ALIASES = (
    "headline",
    "title",
    "見出し",
    "标题",
)
_CREATIVE_BODY_ALIASES = (
    "body",
    "primary text",
    "本文",
    "正文",
)
_CREATIVE_CTA_ALIASES = (
    "call to action",
    "cta",
    "call_to_action",
    "行動を促すフレーズ",
)

# Sentinel values used by Meta to indicate "no breakdown for this row"
# (i.e. a totals row, not a demographic-level detail). Multilingual.
_ALL_SENTINELS = frozenset(
    {
        "all",
        "all (default)",
        "総計",
        "全部",
        "全部 (默认)",
        "전체",
        "todas",
        "tous",
        "alle",
    }
)


_DATE_RE_DASH = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")
_DATE_RE_SLASH_ISO = re.compile(r"^(\d{4})/(\d{2})/(\d{2})$")
_DATE_RE_SLASH_US = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")


def _parse_day(value: str) -> str:
    """Normalize a date cell to YYYY-MM-DD; return '' on failure.

    Recognized formats:
      - ``YYYY-MM-DD`` (ISO)
      - ``YYYY/MM/DD`` (slash-form ISO)
      - ``MM/DD/YYYY`` (US-locale Ads Manager export)
    EU-locale ``DD/MM/YYYY`` is intentionally **not** recognized: it is
    indistinguishable from ``MM/DD/YYYY`` for days <= 12, so accepting
    both would silently mis-aggregate metrics half the time. EU-locale
    users are instructed in ``docs/byod.md`` to switch *Reports →
    Account language* to English (which produces ``MM/DD/YYYY``).
    """
    s = (value or "").strip()
    # Excel/Sheets sometimes serialize date cells as
    # "YYYY-MM-DD HH:MM:SS"; drop the time portion before regex so the
    # cell matches the same way as a bare ISO date string.
    if " " in s:
        s = s.split(" ", 1)[0]
    m = _DATE_RE_DASH.match(s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = _DATE_RE_SLASH_ISO.match(s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = _DATE_RE_SLASH_US.match(s)
    if m:
        return f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return ""


def _synthetic_id(prefix: str, name: str) -> str:
    """Deterministic short ID from a name (stable across imports)."""
    h = hashlib.sha256(name.encode("utf-8")).hexdigest()[:10]
    return f"{prefix}_{h}"


def _header_row(sheet: object) -> list[str]:
    """Return the first row of an openpyxl sheet, lowercased + stripped."""
    rows = list(sheet.iter_rows(values_only=True, max_row=1))  # type: ignore[attr-defined]
    if not rows:
        return []
    return [str(c).strip().lower() if c is not None else "" for c in rows[0]]


def _resolve_alias(header: list[str], aliases: tuple[str, ...]) -> int | None:
    """Return the column index of the first matching alias, or None."""
    for alias in aliases:
        if alias in header:
            return header.index(alias)
    return None


def _iter_data_rows(sheet: object) -> Any:
    """Yield rows after the header, skipping fully-blank lines."""
    first = True
    for row in sheet.iter_rows(values_only=True):  # type: ignore[attr-defined]
        if first:
            first = False
            continue
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        yield row


def _cell_at(row: tuple[Any, ...], idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    return "" if val is None else str(val).strip()


def _detect_meta_sheet(workbook: Workbook) -> str | None:
    """Return the name of the first sheet that looks like a Meta export.

    The Google Ads Script tabs use the column ``campaign`` (short
    form), while every Meta Ads Manager export — across all locales —
    uses a long-form campaign-name header (``Campaign name``,
    ``キャンペーン名``, ``Nombre de la campaña``, …). Requiring any of
    those long-form aliases keeps the Google Ads / Meta adapters
    disjoint when both data sources are bundled in a single workbook,
    while still accepting any localized Meta export.
    """
    for sheet_name in workbook.sheetnames:
        header = _header_row(workbook[sheet_name])
        if not header:
            continue
        has_date = any(a in header for a in _DATE_ALIASES)
        has_campaign = any(a in header for a in _CAMPAIGN_ALIASES)
        has_impressions = any(a in header for a in _IMPRESSIONS_ALIASES)
        if has_date and has_campaign and has_impressions:
            return str(sheet_name)
    return None


@dataclass
class MetaAdsAdapter:
    """Workbook-aware Meta Ads BYOD adapter."""

    @classmethod
    def has_tab(cls, workbook: Workbook) -> bool:
        """True when the workbook contains a sheet whose header looks
        like a Meta Ads Manager export."""
        return _detect_meta_sheet(workbook) is not None

    def normalize_from_workbook(
        self, workbook: Workbook, dst_dir: Path
    ) -> ImportResult:
        sheet_name = _detect_meta_sheet(workbook)
        if sheet_name is None:
            raise UnsupportedFormatError(
                "No sheet matched the Meta Ads export schema. Expected a "
                "tab with at least: a date column (e.g. Day / Reporting "
                "starts / 日 / レポート開始日 / Día / 日期), a campaign "
                "name column (Campaign name / キャンペーン名 / Nombre de "
                "la campaña / 廣告活動名稱 / 广告系列名称 / …), and an "
                "Impressions column (Impressions / インプレッション / "
                "Impresiones / 노출 / 曝光次數 / …)."
            )

        sheet = workbook[sheet_name]
        header = _header_row(sheet)

        date_idx = _resolve_alias(header, _DATE_ALIASES)
        camp_idx = _resolve_alias(header, _CAMPAIGN_ALIASES)
        impr_idx = _resolve_alias(header, _IMPRESSIONS_ALIASES)
        clicks_idx = _resolve_alias(header, _CLICKS_ALIASES)
        spend_idx = _resolve_alias(header, _SPEND_ALIASES)
        conv_idx = _resolve_alias(header, _CONVERSIONS_ALIASES)
        ad_set_idx = _resolve_alias(header, _AD_SET_ALIASES)
        ad_idx = _resolve_alias(header, _AD_ALIASES)
        # Phase 3 — additional analytical columns (all optional). Each
        # falls back to ``None`` when the export does not include the
        # column, in which case the corresponding metric defaults to 0
        # and the corresponding output CSV is suppressed when empty.
        reach_idx = _resolve_alias(header, _REACH_ALIASES)
        freq_idx = _resolve_alias(header, _FREQUENCY_ALIASES)
        ri_idx = _resolve_alias(header, _RESULT_INDICATOR_ALIASES)
        age_idx = _resolve_alias(header, _AGE_ALIASES)
        gender_idx = _resolve_alias(header, _GENDER_ALIASES)
        region_idx = _resolve_alias(header, _REGION_ALIASES)
        placement_idx = _resolve_alias(header, _PLACEMENT_ALIASES)
        platform_idx = _resolve_alias(header, _PLATFORM_ALIASES)
        device_idx = _resolve_alias(header, _DEVICE_ALIASES)
        img_idx = _resolve_alias(header, _CREATIVE_IMAGE_URL_ALIASES)
        vid_idx = _resolve_alias(header, _CREATIVE_VIDEO_URL_ALIASES)
        headline_idx = _resolve_alias(header, _CREATIVE_HEADLINE_ALIASES)
        body_idx = _resolve_alias(header, _CREATIVE_BODY_ALIASES)
        cta_idx = _resolve_alias(header, _CREATIVE_CTA_ALIASES)

        if date_idx is None or camp_idx is None or impr_idx is None:
            raise UnsupportedFormatError(
                f"{sheet_name}: missing required columns. Need a date "
                f"column, a campaign-name column, and an impressions "
                f"column (any supported locale). Found: {header}"
            )

        dst_dir.mkdir(parents=True, exist_ok=True)
        files_written: list[str] = []
        all_dates: list[str] = []

        # Aggregate across rows. The export may have multiple rows per
        # (day, campaign) pair when Ad set / Ad breakdown is enabled,
        # so metrics need summing per day×campaign rather than
        # passthrough.
        campaign_ids: dict[str, str] = {}
        ad_set_ids: dict[tuple[str, str], str] = {}  # (campaign, ad_set) -> id
        # Keying ads by (campaign, ad_set, ad) — not (ad_set, ad) — so
        # that two campaigns reusing the same ad set name (e.g.
        # "Default", "Lookalike 1%") get distinct ad_set_id rows in
        # ads.csv. Earlier `(ad_set, ad)` keying caused a cross-campaign
        # collision flagged in code review of this PR.
        ad_records: list[tuple[str, str, str, str]] = []  # (camp, ad_set, ad, id)
        # O(1) lookup keyed by (camp, ad_set, ad) — the previous list
        # scan was O(N²) in the row count of the workbook.
        ad_ids_dict: dict[tuple[str, str, str], str] = {}

        # day×campaign → metrics (campaign-level rollup, the primary
        # daily-check input). Stores 0 for any metric whose source
        # column is absent from the workbook.
        metrics_agg: dict[tuple[str, str], dict[str, float]] = {}
        # Phase 3-1 — first non-empty result_indicator value seen per
        # (day, campaign), so daily-check can tell what the "Results"
        # column actually counts (e.g. lead, purchase, link_click).
        metrics_result_indicator: dict[tuple[str, str], str] = {}
        # Phase 3-2 — finer-grain aggregation. Populated only when the
        # export carries Ad set / Ad columns; otherwise the dicts stay
        # empty and the corresponding CSVs are suppressed.
        ad_set_metrics_agg: dict[tuple[str, str, str], dict[str, float]] = {}
        ad_metrics_agg: dict[tuple[str, str, str, str], dict[str, float]] = {}
        # Phase 3-3 — demographics breakdown rows. Activated only when
        # at least one row has a non-"All" value in age/gender/region/
        # placement. Subtotal rows ("All" / 全部 / etc.) are filtered
        # out so totals from the breakdown rows are not double-counted.
        demo_agg: dict[tuple[str, str, str, str], dict[str, float]] = {}
        # Phase 3-4 — creatives by ad_id. Best-effort: only populated
        # when the export carries image_url / video_url / headline /
        # body / cta columns. Empty strings allowed per field.
        creatives: dict[str, dict[str, str]] = {}

        # Currency validation is interleaved with the main loop below
        # rather than pre-scanned, because openpyxl's ``read_only=True``
        # sheets are effectively single-pass — a separate scan loop
        # would consume the row iterator and leave the main loop with
        # no rows to process.

        # Pivot exports often carry BOTH a campaign-level rollup row
        # ("Tokyo"/"All", "Video A"/"All", "All", ...) AND ad-level
        # detail rows for the same (day, campaign). Summing both into
        # metrics_agg would double-count. Defensive rule: pre-buffer
        # rows, determine the deepest grain present for each (day,
        # campaign), and only aggregate rows at that grain. Demographic
        # rows are routed separately and are exempt from this dedup.
        buffered_rows: list[tuple[Any, ...]] = list(_iter_data_rows(sheet))

        # Pass 1 — discover the deepest non-demographic grain per
        # (day, campaign). 0 = campaign rollup, 1 = ad-set, 2 = ad.
        # Also track which (day, campaign) pairs have ANY non-demographic
        # row, so that an export carrying ONLY breakdown rows
        # (e.g. placement × platform × device) still feeds metrics_daily
        # by summing the breakdown rows up to the campaign×day grain.
        deepest_grain: dict[tuple[str, str], int] = {}
        has_non_demo: set[tuple[str, str]] = set()
        breakdown_indices = (
            age_idx,
            gender_idx,
            region_idx,
            placement_idx,
            platform_idx,
            device_idx,
        )
        for raw in buffered_rows:
            day_p1 = _parse_day(_cell_at(raw, date_idx))
            if not day_p1:
                continue
            camp_p1 = _cell_at(raw, camp_idx)
            if not camp_p1:
                continue
            cid_p1 = campaign_ids.setdefault(camp_p1, _synthetic_id("camp", camp_p1))
            key_p1 = (day_p1, cid_p1)
            is_demo_p1 = False
            for idx_p1 in breakdown_indices:
                v = _cell_at(raw, idx_p1)
                if v and v.lower() not in _ALL_SENTINELS:
                    is_demo_p1 = True
                    break
            if is_demo_p1:
                continue
            has_non_demo.add(key_p1)
            as_p1 = _cell_at(raw, ad_set_idx)
            ad_p1 = _cell_at(raw, ad_idx)
            as_p1 = "" if not as_p1 or as_p1.lower() in _ALL_SENTINELS else as_p1
            ad_p1 = "" if not ad_p1 or ad_p1.lower() in _ALL_SENTINELS else ad_p1
            grain = 2 if (as_p1 and ad_p1) else (1 if as_p1 else 0)
            if grain > deepest_grain.get(key_p1, -1):
                deepest_grain[key_p1] = grain

        # Pass 2 — main aggregation. Skip non-leaf rollup rows whose
        # grain is shallower than the deepest detail row we saw above.
        for raw in buffered_rows:
            day = _parse_day(_cell_at(raw, date_idx))
            if not day:
                continue
            camp_name = _cell_at(raw, camp_idx)
            if not camp_name:
                continue

            cid = campaign_ids.setdefault(camp_name, _synthetic_id("camp", camp_name))

            # Treat "All" / locale equivalent in the ad-set / ad column
            # as "no breakdown for this row" — i.e., the row is a
            # campaign-level rollup that should not feed ad_set / ad
            # CSVs (otherwise pivot exports would double-count).
            ad_set_name_raw = _cell_at(raw, ad_set_idx)
            ad_set_name = (
                ad_set_name_raw
                if ad_set_name_raw and ad_set_name_raw.lower() not in _ALL_SENTINELS
                else ""
            )
            asid = ""
            if ad_set_name:
                key = (camp_name, ad_set_name)
                if key not in ad_set_ids:
                    ad_set_ids[key] = _synthetic_id("as", f"{camp_name}::{ad_set_name}")
                asid = ad_set_ids[key]

            ad_name_raw = _cell_at(raw, ad_idx)
            ad_name = (
                ad_name_raw
                if ad_name_raw and ad_name_raw.lower() not in _ALL_SENTINELS
                else ""
            )
            aid = ""
            if ad_name and ad_set_name:
                ad_key = (camp_name, ad_set_name, ad_name)
                aid = ad_ids_dict.get(ad_key, "")
                if not aid:
                    aid = _synthetic_id("ad", f"{camp_name}::{ad_set_name}::{ad_name}")
                    ad_ids_dict[ad_key] = aid
                    ad_records.append((camp_name, ad_set_name, ad_name, aid))

            spend_raw = _cell_at(raw, spend_idx)
            if spend_raw and spend_raw[0] in _NON_JPY_CURRENCY_PREFIXES:
                raise UnsupportedFormatError(
                    f"{sheet_name}: spend column contains non-JPY value "
                    f"{spend_raw!r}. The BYOD pipeline assumes JPY; "
                    f"switch Ads Manager → Account currency to JPY "
                    f"before export."
                )

            impressions = _to_float(_cell_at(raw, impr_idx))
            clicks = _to_float(_cell_at(raw, clicks_idx))
            cost = _to_float(spend_raw)
            conv = _to_float(_cell_at(raw, conv_idx))
            reach = _to_float(_cell_at(raw, reach_idx))
            # The export's Frequency column (when present) is read
            # only for completeness — the value written to
            # metrics_daily.csv is always derived from the aggregated
            # impressions/reach to keep the math correct under
            # multi-row aggregation.
            result_indicator = _cell_at(raw, ri_idx)

            # Detect demographic breakdown — first non-"All" value across
            # the supported dimensions wins the dimension label. Multi-
            # dimensional breakdown rows (e.g. placement × platform ×
            # device, all set on one row) are recorded under the first
            # matched dimension only; finer granularity is lost in v1.
            demo_dim: str = ""
            demo_val: str = ""
            for dim, idx in (
                ("age", age_idx),
                ("gender", gender_idx),
                ("region", region_idx),
                ("placement", placement_idx),
                ("platform", platform_idx),
                ("device", device_idx),
            ):
                val = _cell_at(raw, idx)
                if val and val.lower() not in _ALL_SENTINELS:
                    demo_dim, demo_val = dim, val
                    break

            # Skip non-leaf rollup rows when finer-grain detail rows
            # exist for the same (day, campaign). Without this guard
            # a pivot export carrying both a campaign rollup row and
            # ad-level detail rows for the same day×campaign would
            # double-count metrics_agg (rollup + sum(details)).
            if not demo_dim:
                row_grain = 2 if ad_name else (1 if ad_set_name else 0)
                if row_grain < deepest_grain.get((day, cid), 0):
                    all_dates.append(day)
                    continue

            if demo_dim:
                # Demographic-breakdown row → demographics_daily.csv.
                d_key = (day, cid, demo_dim, demo_val)
                d = demo_agg.setdefault(
                    d_key,
                    {
                        "impressions": 0.0,
                        "clicks": 0.0,
                        "cost": 0.0,
                        "conv": 0.0,
                        "reach": 0.0,
                    },
                )
                d["impressions"] += impressions
                d["clicks"] += clicks
                d["cost"] += cost
                d["conv"] += conv
                d["reach"] += reach

                # Fallback: when the workbook has NO non-demographic
                # row for this (day, campaign), the breakdown rows are
                # the only signal we have — sum them up to the
                # campaign×day grain so metrics_daily isn't empty.
                # Caveat: reach is unique users per breakdown cell;
                # summing across cells over-counts users seen on
                # multiple cells. This is an acceptable approximation
                # for the v1 BYOD path. (The export with both rollup
                # AND breakdown rows takes the rollup-row branch above
                # via has_non_demo and skips this fallback.)
                if (day, cid) not in has_non_demo:
                    fb = metrics_agg.setdefault(
                        (day, cid),
                        {
                            "impressions": 0.0,
                            "clicks": 0.0,
                            "cost": 0.0,
                            "conv": 0.0,
                            "reach": 0.0,
                        },
                    )
                    fb["impressions"] += impressions
                    fb["clicks"] += clicks
                    fb["cost"] += cost
                    fb["conv"] += conv
                    fb["reach"] += reach
                    if result_indicator and (day, cid) not in metrics_result_indicator:
                        metrics_result_indicator[(day, cid)] = result_indicator

                    # The same row also identifies an ad-set / ad when
                    # those columns are populated. Aggregate to those
                    # grains so drill-down CSVs (ad_set_metrics_daily,
                    # ad_metrics_daily) are written even when every row
                    # carries a breakdown dimension. Each output is a
                    # different aggregation view; the same row legally
                    # contributes to multiple views.
                    if asid:
                        as_fb = ad_set_metrics_agg.setdefault(
                            (day, cid, asid),
                            {
                                "impressions": 0.0,
                                "clicks": 0.0,
                                "cost": 0.0,
                                "conv": 0.0,
                                "reach": 0.0,
                            },
                        )
                        as_fb["impressions"] += impressions
                        as_fb["clicks"] += clicks
                        as_fb["cost"] += cost
                        as_fb["conv"] += conv
                        as_fb["reach"] += reach

                    if aid:
                        ad_fb = ad_metrics_agg.setdefault(
                            (day, cid, asid, aid),
                            {
                                "impressions": 0.0,
                                "clicks": 0.0,
                                "cost": 0.0,
                                "conv": 0.0,
                                "reach": 0.0,
                            },
                        )
                        ad_fb["impressions"] += impressions
                        ad_fb["clicks"] += clicks
                        ad_fb["cost"] += cost
                        ad_fb["conv"] += conv
                        ad_fb["reach"] += reach
            else:
                agg_key = (day, cid)
                cell = metrics_agg.setdefault(
                    agg_key,
                    {
                        "impressions": 0.0,
                        "clicks": 0.0,
                        "cost": 0.0,
                        "conv": 0.0,
                        "reach": 0.0,
                    },
                )
                cell["impressions"] += impressions
                cell["clicks"] += clicks
                cell["cost"] += cost
                cell["conv"] += conv
                cell["reach"] += reach
                if result_indicator and agg_key not in metrics_result_indicator:
                    metrics_result_indicator[agg_key] = result_indicator

                if asid:
                    as_key = (day, cid, asid)
                    a = ad_set_metrics_agg.setdefault(
                        as_key,
                        {
                            "impressions": 0.0,
                            "clicks": 0.0,
                            "cost": 0.0,
                            "conv": 0.0,
                            "reach": 0.0,
                        },
                    )
                    a["impressions"] += impressions
                    a["clicks"] += clicks
                    a["cost"] += cost
                    a["conv"] += conv
                    a["reach"] += reach

                if aid:
                    ad_metrics_key = (day, cid, asid, aid)
                    ad_cell = ad_metrics_agg.setdefault(
                        ad_metrics_key,
                        {
                            "impressions": 0.0,
                            "clicks": 0.0,
                            "cost": 0.0,
                            "conv": 0.0,
                            "reach": 0.0,
                        },
                    )
                    ad_cell["impressions"] += impressions
                    ad_cell["clicks"] += clicks
                    ad_cell["cost"] += cost
                    ad_cell["conv"] += conv
                    ad_cell["reach"] += reach

            # Creative info — captured per ad_id, but only when at
            # least one creative field on this row is non-empty. This
            # prevents a later row with all-empty creative cells from
            # clobbering an earlier row that carried the URLs / copy.
            if aid and any(
                idx is not None
                for idx in (img_idx, vid_idx, headline_idx, body_idx, cta_idx)
            ):
                creative_row = {
                    "image_url": _sanitize_cell(_cell_at(raw, img_idx)),
                    "video_url": _sanitize_cell(_cell_at(raw, vid_idx)),
                    "headline": _sanitize_cell(_cell_at(raw, headline_idx)),
                    "body": _sanitize_cell(_cell_at(raw, body_idx)),
                    "cta": _sanitize_cell(_cell_at(raw, cta_idx)),
                }
                if any(creative_row.values()):
                    creatives[aid] = {
                        "ad_id": aid,
                        "name": _sanitize_cell(ad_name),
                        **creative_row,
                    }

            all_dates.append(day)

        if not metrics_agg and not demo_agg:
            raise UnsupportedFormatError(
                f"{sheet_name}: no data rows after the header."
            )

        # ---- campaigns.csv ---------------------------------------------------
        campaigns_path = dst_dir / "campaigns.csv"
        with campaigns_path.open("w", encoding="utf-8", newline="") as f:
            writer = _csv.DictWriter(
                f,
                fieldnames=[
                    "campaign_id",
                    "name",
                    "status",
                    "objective",
                    "daily_budget_jpy",
                ],
            )
            writer.writeheader()
            for name, cid in campaign_ids.items():
                writer.writerow(
                    {
                        "campaign_id": cid,
                        "name": _sanitize_cell(name),
                        # Ads Manager export does not carry status /
                        # objective / budget. Empty strings keep the
                        # column shape stable for the BYOD client's
                        # _to_float / _to_int helpers.
                        "status": "",
                        "objective": "",
                        "daily_budget_jpy": "",
                    }
                )
        files_written.append("campaigns.csv")

        # ---- ad_sets.csv -----------------------------------------------------
        if ad_set_ids:
            ad_sets_path = dst_dir / "ad_sets.csv"
            with ad_sets_path.open("w", encoding="utf-8", newline="") as f:
                writer = _csv.DictWriter(
                    f,
                    fieldnames=["ad_set_id", "campaign_id", "name", "status"],
                )
                writer.writeheader()
                for (camp_name, ad_set_name), asid in ad_set_ids.items():
                    writer.writerow(
                        {
                            "ad_set_id": asid,
                            "campaign_id": campaign_ids[camp_name],
                            "name": _sanitize_cell(ad_set_name),
                            "status": "",
                        }
                    )
            files_written.append("ad_sets.csv")

        # ---- ads.csv ---------------------------------------------------------
        if ad_records:
            ads_path = dst_dir / "ads.csv"
            with ads_path.open("w", encoding="utf-8", newline="") as f:
                writer = _csv.DictWriter(
                    f,
                    fieldnames=["ad_id", "ad_set_id", "name", "status"],
                )
                writer.writeheader()
                for camp_name, ad_set_name, ad_name, aid in ad_records:
                    # Direct lookup keyed by (campaign, ad_set) — fixes
                    # the cross-campaign collision flagged in review.
                    asid = ad_set_ids.get((camp_name, ad_set_name), "")
                    writer.writerow(
                        {
                            "ad_id": aid,
                            "ad_set_id": asid,
                            "name": _sanitize_cell(ad_name),
                            "status": "",
                        }
                    )
            files_written.append("ads.csv")

        # ---- metrics_daily.csv ----------------------------------------------
        metrics_path = dst_dir / "metrics_daily.csv"
        with metrics_path.open("w", encoding="utf-8", newline="") as f:
            writer = _csv.DictWriter(
                f,
                fieldnames=[
                    "date",
                    "campaign_id",
                    "impressions",
                    "clicks",
                    "cost_jpy",
                    "conversions",
                    # Phase 3-1 — extra analytical columns. Empty
                    # strings preserve column shape for clients that
                    # tolerate missing data via _to_float / _to_int.
                    "reach",
                    "frequency",
                    "result_indicator",
                ],
            )
            writer.writeheader()
            for (day, cid), m in sorted(metrics_agg.items()):
                # Frequency is impressions/reach. Always derive it
                # from the aggregated totals instead of averaging the
                # per-row Frequency column — averaging an arithmetic
                # mean of ratios would understate frequency for
                # high-impression rows.
                freq_value = (
                    round(m["impressions"] / m["reach"], 2) if m["reach"] > 0 else 0.0
                )
                writer.writerow(
                    {
                        "date": day,
                        "campaign_id": cid,
                        "impressions": int(m["impressions"]),
                        "clicks": int(m["clicks"]),
                        "cost_jpy": _round2(m["cost"]),
                        "conversions": _round2(m["conv"]),
                        "reach": int(m["reach"]),
                        "frequency": freq_value,
                        "result_indicator": metrics_result_indicator.get(
                            (day, cid), ""
                        ),
                    }
                )
        files_written.append("metrics_daily.csv")

        # ---- ad_set_metrics_daily.csv (Phase 3-2) --------------------------
        if ad_set_metrics_agg:
            as_metrics_path = dst_dir / "ad_set_metrics_daily.csv"
            with as_metrics_path.open("w", encoding="utf-8", newline="") as f:
                writer = _csv.DictWriter(
                    f,
                    fieldnames=[
                        "date",
                        "campaign_id",
                        "ad_set_id",
                        "impressions",
                        "clicks",
                        "cost_jpy",
                        "conversions",
                        "reach",
                    ],
                )
                writer.writeheader()
                for (day, cid, asid), m in sorted(ad_set_metrics_agg.items()):
                    writer.writerow(
                        {
                            "date": day,
                            "campaign_id": cid,
                            "ad_set_id": asid,
                            "impressions": int(m["impressions"]),
                            "clicks": int(m["clicks"]),
                            "cost_jpy": _round2(m["cost"]),
                            "conversions": _round2(m["conv"]),
                            "reach": int(m["reach"]),
                        }
                    )
            files_written.append("ad_set_metrics_daily.csv")

        # ---- ad_metrics_daily.csv (Phase 3-2) -------------------------------
        if ad_metrics_agg:
            ad_metrics_path = dst_dir / "ad_metrics_daily.csv"
            with ad_metrics_path.open("w", encoding="utf-8", newline="") as f:
                writer = _csv.DictWriter(
                    f,
                    fieldnames=[
                        "date",
                        "campaign_id",
                        "ad_set_id",
                        "ad_id",
                        "impressions",
                        "clicks",
                        "cost_jpy",
                        "conversions",
                        "reach",
                    ],
                )
                writer.writeheader()
                for (day, cid, asid, aid), m in sorted(ad_metrics_agg.items()):
                    writer.writerow(
                        {
                            "date": day,
                            "campaign_id": cid,
                            "ad_set_id": asid,
                            "ad_id": aid,
                            "impressions": int(m["impressions"]),
                            "clicks": int(m["clicks"]),
                            "cost_jpy": _round2(m["cost"]),
                            "conversions": _round2(m["conv"]),
                            "reach": int(m["reach"]),
                        }
                    )
            files_written.append("ad_metrics_daily.csv")

        # ---- demographics_daily.csv (Phase 3-3) -----------------------------
        if demo_agg:
            demo_path = dst_dir / "demographics_daily.csv"
            with demo_path.open("w", encoding="utf-8", newline="") as f:
                writer = _csv.DictWriter(
                    f,
                    fieldnames=[
                        "date",
                        "campaign_id",
                        "dimension",
                        "value",
                        "impressions",
                        "clicks",
                        "cost_jpy",
                        "conversions",
                        "reach",
                    ],
                )
                writer.writeheader()
                for (day, cid, dim, val), d in sorted(demo_agg.items()):
                    writer.writerow(
                        {
                            "date": day,
                            "campaign_id": cid,
                            "dimension": dim,
                            "value": _sanitize_cell(val),
                            "impressions": int(d["impressions"]),
                            "clicks": int(d["clicks"]),
                            "cost_jpy": _round2(d["cost"]),
                            "conversions": _round2(d["conv"]),
                            "reach": int(d["reach"]),
                        }
                    )
            files_written.append("demographics_daily.csv")

        # ---- creatives.csv (Phase 3-4) --------------------------------------
        if creatives:
            creatives_path = dst_dir / "creatives.csv"
            with creatives_path.open("w", encoding="utf-8", newline="") as f:
                writer = _csv.DictWriter(
                    f,
                    fieldnames=[
                        "ad_id",
                        "name",
                        "image_url",
                        "video_url",
                        "headline",
                        "body",
                        "cta",
                    ],
                )
                writer.writeheader()
                for c in creatives.values():
                    writer.writerow(c)
            files_written.append("creatives.csv")

        sorted_dates = sorted(all_dates)
        return ImportResult(
            rows=len(metrics_agg),
            date_range=(sorted_dates[0], sorted_dates[-1]),
            files_written=files_written,
            source_format=SOURCE_FORMAT,
            campaigns=len(campaign_ids),
            # ImportResult.ad_groups is repurposed for Meta as ad-set
            # count for parity with the Google Ads adapter; the
            # `ad_groups` manifest key documents downstream as
            # "second-level identity rows" rather than the literal
            # Google Ads ad-group concept.
            ad_groups=len(ad_set_ids),
        )


def _to_float(value: str) -> float:
    """Parse a numeric cell that may include a currency symbol or
    thousands separators (Ads Manager export sometimes wraps ``Amount
    spent`` as ``¥1,234.56`` or ``"1,234"``). Returns 0.0 on failure.

    Only ¥ (JPY) and bare numerics are tolerated. Non-JPY currency
    symbols (``$``, ``€``, ``£``) leak through as 0.0 and are caught
    upstream by :func:`_scan_non_jpy_currency` before this function
    is reached on a real spend cell.
    """
    s = (value or "").strip()
    if not s:
        return 0.0
    s = s.replace(",", "").replace("¥", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _round2(value: float) -> str:
    return f"{value:.2f}"


_NON_JPY_CURRENCY_PREFIXES = ("$", "€", "£", "₩", "₹", "¢")


# Cells starting with one of these characters are treated as formulas
# by Excel / Google Sheets when the CSV is re-opened. A campaign named
# ``=cmd|...`` would auto-execute on re-open, exfiltrating data. We
# sanitize untrusted cell values by prefixing a single quote.
# OWASP "CSV Injection" — the leading quote is stripped on display by
# Excel and renders as a literal at the start of the field elsewhere.
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_cell(value: str) -> str:
    """Defang user-controlled cell content against CSV-injection."""
    if value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value
