"""Google Ads BYOD adapter — consumes the Ads tabs of the mureo Sheet bundle.

The Google Ads side of the mureo Sheet template is populated by
``scripts/sheet-template/google-ads-script.js`` running inside the
user's Google Ads UI (Tools → Bulk actions → Scripts). It writes five
fixed-schema tabs to the bound Sheet:

  campaigns         day, campaign, impressions, clicks, cost, conversions
  ad_groups         day, campaign, ad_group, impressions, clicks, cost, conversions
  search_terms      search_term, campaign, ad_group, impressions, clicks, cost, conversions
  keywords          keyword, match_type, quality_score, campaign, ad_group,
                    impressions, clicks, cost, conversions
  auction_insights  campaign, competitor_domain, impression_share, outranking_share

The adapter normalizes those tabs to the CSV layout the existing BYOD
Google Ads client (``mureo/byod/clients.py``) reads under
``~/.mureo/byod/google_ads/``:

  campaigns.csv        — campaign-level identity + state placeholders
  ad_groups.csv        — ad-group-level identity, parent campaign_id
  metrics_daily.csv    — day × campaign rollup
  keywords.csv         — keyword-level (Apps Script tab passthrough)
  search_terms.csv     — search-term-level (passthrough)
  auction_insights.csv — competitor share (passthrough)

Campaign / ad-group identity is synthesized from name (deterministic
hash) because the Apps Script does not export numeric IDs. The hash is
stable across runs for the same name, so re-imports keep the same
``campaign_id`` values and downstream STATE.json references continue
to resolve.
"""

from __future__ import annotations

import csv as _csv
import hashlib
import re
from dataclasses import dataclass
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.workbook.workbook import Workbook

SOURCE_FORMAT = "mureo_sheet_bundle_google_ads_v1"

CAMPAIGNS_TAB = "campaigns"
AD_GROUPS_TAB = "ad_groups"
SEARCH_TERMS_TAB = "search_terms"
KEYWORDS_TAB = "keywords"
AUCTION_INSIGHTS_TAB = "auction_insights"


@dataclass
class ImportResult:
    """Per-platform return value from a normalize_from_workbook call."""

    rows: int
    date_range: tuple[str, str]
    files_written: list[str]
    source_format: str
    campaigns: int
    ad_groups: int


class UnsupportedFormatError(ValueError):
    """Raised when a workbook tab violates its required schema."""


_DATE_RE_DASH = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")
_DATE_RE_SLASH = re.compile(r"^(\d{4})/(\d{2})/(\d{2})$")


def _parse_day(value: str) -> str:
    """Normalize a date cell to YYYY-MM-DD; return '' on failure."""
    s = (value or "").strip()
    m = _DATE_RE_DASH.match(s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = _DATE_RE_SLASH.match(s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return ""


def _synthetic_id(prefix: str, name: str) -> str:
    """Deterministic short ID from a name (stable across imports)."""
    h = hashlib.sha256(name.encode("utf-8")).hexdigest()[:10]
    return f"{prefix}_{h}"


def _header_row(sheet: object) -> list[str]:
    """Return the first row of an openpyxl sheet, lowercased + stripped."""
    rows = list(sheet.iter_rows(values_only=True, max_row=1))  # type: ignore[attr-defined]
    if not rows:
        return []
    return [str(c).strip().lower() if c is not None else "" for c in rows[0]]


def _require_columns(
    sheet: object, tab: str, required: tuple[str, ...]
) -> dict[str, int]:
    header = _header_row(sheet)
    for col in required:
        if col not in header:
            raise UnsupportedFormatError(
                f"{tab}: missing required column {col!r}. Found: {header}"
            )
    return {c: header.index(c) for c in header if c}


def _iter_data_rows(sheet: object) -> Any:
    """Yield rows after the header, skipping fully-blank lines."""
    first = True
    for row in sheet.iter_rows(values_only=True):  # type: ignore[attr-defined]
        if first:
            first = False
            continue
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        yield row


def _cell(row: tuple[Any, ...], idx: dict[str, int], col: str) -> str:
    if col not in idx:
        return ""
    pos = idx[col]
    if pos >= len(row):
        return ""
    val = row[pos]
    return "" if val is None else str(val).strip()


@dataclass
class GoogleAdsAdapter:
    """Workbook-aware Google Ads bundle adapter."""

    @classmethod
    def has_tab(cls, workbook: Workbook) -> bool:
        """True when at least the ``campaigns`` tab is present.

        Other tabs are optional — running the Sheet's Apps Script
        without a Google Ads account, or running the Ads Script with
        only some campaign types accessible, may legitimately leave
        ``ad_groups`` / ``search_terms`` / ``keywords`` /
        ``auction_insights`` absent.
        """
        return CAMPAIGNS_TAB in workbook.sheetnames

    def normalize_from_workbook(
        self, workbook: Workbook, dst_dir: Path
    ) -> ImportResult:
        if CAMPAIGNS_TAB not in workbook.sheetnames:
            raise UnsupportedFormatError(
                f"Required tab {CAMPAIGNS_TAB!r} missing from the workbook."
            )

        dst_dir.mkdir(parents=True, exist_ok=True)
        files_written: list[str] = []
        all_dates: list[str] = []

        # ---- campaigns tab → campaigns.csv + metrics_daily.csv -----
        camp_sheet = workbook[CAMPAIGNS_TAB]
        camp_idx = _require_columns(
            camp_sheet,
            CAMPAIGNS_TAB,
            ("day", "campaign", "impressions", "clicks", "cost"),
        )
        campaign_ids: dict[str, str] = {}
        metrics_rows: list[dict[str, str]] = []
        for raw in _iter_data_rows(camp_sheet):
            day = _parse_day(_cell(raw, camp_idx, "day"))
            if not day:
                continue
            name = _cell(raw, camp_idx, "campaign")
            if not name:
                continue
            cid = campaign_ids.setdefault(name, _synthetic_id("camp", name))
            # Output column name is cost_jpy (not cost) so the existing
            # ByodGoogleAdsClient — which reads r.get("cost_jpy") in
            # mureo/byod/clients.py — sees real spend instead of 0.
            metrics_rows.append(
                {
                    "date": day,
                    "campaign_id": cid,
                    "impressions": _cell(raw, camp_idx, "impressions"),
                    "clicks": _cell(raw, camp_idx, "clicks"),
                    "cost_jpy": _cell(raw, camp_idx, "cost"),
                    "conversions": _cell(raw, camp_idx, "conversions"),
                }
            )
            all_dates.append(day)

        if not metrics_rows:
            raise UnsupportedFormatError(
                f"{CAMPAIGNS_TAB}: no data rows after the header."
            )

        # campaigns.csv — identity table for the BYOD client. Column
        # names match what mureo/byod/clients.py:_campaign_to_dict
        # reads (name, status, daily_budget_jpy etc.) so the existing
        # ByodGoogleAdsClient consumes our output unchanged.
        campaigns_path = dst_dir / "campaigns.csv"
        with campaigns_path.open("w", encoding="utf-8", newline="") as f:
            writer = _csv.DictWriter(
                f,
                fieldnames=[
                    "campaign_id",
                    "name",
                    "status",
                    "daily_budget_jpy",
                ],
            )
            writer.writeheader()
            for name, cid in campaign_ids.items():
                writer.writerow(
                    {
                        "campaign_id": cid,
                        "name": name,
                        # The Apps Script tab does not export state / budget.
                        # Empty strings keep the CSV column-stable; the
                        # client's _to_float / _to_int helpers tolerate them.
                        "status": "",
                        "daily_budget_jpy": "",
                    }
                )
        files_written.append("campaigns.csv")

        # metrics_daily.csv — day × campaign metrics. Column name
        # cost_jpy matches the existing ByodGoogleAdsClient contract;
        # see commentary above the metrics_rows construction.
        metrics_path = dst_dir / "metrics_daily.csv"
        with metrics_path.open("w", encoding="utf-8", newline="") as f:
            writer = _csv.DictWriter(
                f,
                fieldnames=[
                    "date",
                    "campaign_id",
                    "impressions",
                    "clicks",
                    "cost_jpy",
                    "conversions",
                ],
            )
            writer.writeheader()
            for r in metrics_rows:
                writer.writerow(r)
        files_written.append("metrics_daily.csv")

        # ---- ad_groups tab → ad_groups.csv (optional) --------------
        ad_groups_count = 0
        if AD_GROUPS_TAB in workbook.sheetnames:
            ag_sheet = workbook[AD_GROUPS_TAB]
            ag_idx = _require_columns(
                ag_sheet,
                AD_GROUPS_TAB,
                ("campaign", "ad_group"),
            )
            ad_group_ids: dict[tuple[str, str], str] = {}
            for raw in _iter_data_rows(ag_sheet):
                camp_name = _cell(raw, ag_idx, "campaign")
                ag_name = _cell(raw, ag_idx, "ad_group")
                if not camp_name or not ag_name:
                    continue
                key = (camp_name, ag_name)
                if key not in ad_group_ids:
                    ad_group_ids[key] = _synthetic_id("ag", f"{camp_name}::{ag_name}")
                    # Ensure parent campaign exists in the campaign map
                    # even if it had no row in the campaigns tab (rare).
                    campaign_ids.setdefault(camp_name, _synthetic_id("camp", camp_name))

            ad_groups_path = dst_dir / "ad_groups.csv"
            with ad_groups_path.open("w", encoding="utf-8", newline="") as f:
                writer = _csv.DictWriter(
                    f,
                    fieldnames=["ad_group_id", "campaign_id", "name"],
                )
                writer.writeheader()
                for (camp_name, ag_name), agid in ad_group_ids.items():
                    writer.writerow(
                        {
                            "ad_group_id": agid,
                            "campaign_id": campaign_ids[camp_name],
                            "name": ag_name,
                        }
                    )
            files_written.append("ad_groups.csv")
            ad_groups_count = len(ad_group_ids)

        # ---- keywords tab → keywords.csv (optional) ----------------
        if KEYWORDS_TAB in workbook.sheetnames:
            _passthrough_tab(
                workbook[KEYWORDS_TAB],
                dst_dir / "keywords.csv",
                tab_label=KEYWORDS_TAB,
                required=("keyword", "campaign", "ad_group"),
                optional=(
                    "match_type",
                    "quality_score",
                    "impressions",
                    "clicks",
                    "cost",
                    "conversions",
                ),
            )
            files_written.append("keywords.csv")

        # ---- search_terms tab → search_terms.csv (optional) --------
        if SEARCH_TERMS_TAB in workbook.sheetnames:
            _passthrough_tab(
                workbook[SEARCH_TERMS_TAB],
                dst_dir / "search_terms.csv",
                tab_label=SEARCH_TERMS_TAB,
                required=("search_term", "campaign", "ad_group"),
                optional=(
                    "impressions",
                    "clicks",
                    "cost",
                    "conversions",
                ),
            )
            files_written.append("search_terms.csv")

        # ---- auction_insights tab → auction_insights.csv -----------
        if AUCTION_INSIGHTS_TAB in workbook.sheetnames:
            _passthrough_tab(
                workbook[AUCTION_INSIGHTS_TAB],
                dst_dir / "auction_insights.csv",
                tab_label=AUCTION_INSIGHTS_TAB,
                required=("campaign", "competitor_domain"),
                optional=("impression_share", "outranking_share"),
            )
            files_written.append("auction_insights.csv")

        all_dates.sort()
        return ImportResult(
            rows=len(metrics_rows),
            date_range=(all_dates[0], all_dates[-1]),
            files_written=files_written,
            source_format=SOURCE_FORMAT,
            campaigns=len(campaign_ids),
            ad_groups=ad_groups_count,
        )


def _passthrough_tab(
    sheet: object,
    out_path: Path,
    *,
    tab_label: str,
    required: tuple[str, ...],
    optional: tuple[str, ...],
) -> None:
    """Validate required columns then write the tab to CSV verbatim."""
    idx = _require_columns(sheet, tab_label, required)
    out_columns = [c for c in (*required, *optional) if c in idx]
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = _csv.DictWriter(f, fieldnames=out_columns)
        writer.writeheader()
        for raw in _iter_data_rows(sheet):
            writer.writerow({c: _cell(raw, idx, c) for c in out_columns})
