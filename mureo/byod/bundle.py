"""BYOD Sheet bundle importer — single XLSX in, multi-platform out.

The user runs the mureo Google Ads Script (see
``scripts/sheet-template/``), exports the resulting Sheet as XLSX via
*File → Download → Microsoft Excel*, and feeds it to mureo here.
This module is the single entry point for ingesting that file.

Pipeline:

  ``user-bundle.xlsx``
        │
        ▼
  openpyxl ``Workbook``
        │
        └─ GoogleAdsAdapter →  ``~/.mureo/byod/google_ads/``

GA4 and Search Console BYOD paths were removed in Phase 1 of the
BYOD redesign — those platforms remain accessible via the existing
real-API OAuth path. Meta Ads BYOD will land in a follow-up.

Manifest update is atomic. Either the dispatched platform's section is
written to ``manifest.json`` together with on-disk files, or nothing
is.
"""

from __future__ import annotations

import contextlib
import logging
import shutil
from datetime import datetime, timezone
from typing import TYPE_CHECKING, Any

from mureo.byod.adapters.google_ads import (
    GoogleAdsAdapter,
    UnsupportedFormatError,
)
from mureo.byod.runtime import (
    SCHEMA_VERSION,
    byod_data_dir,
    read_manifest,
    write_manifest,
)

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.workbook.workbook import Workbook

logger = logging.getLogger(__name__)


class BundleImportError(RuntimeError):
    """Raised when a bundle import fails for any reason."""


# Adapters registered for the bundle flow. Each adapter's ``has_tab``
# is checked, and the adapter is dispatched only when its required tab
# is present.
_BUNDLE_ADAPTERS: list[tuple[str, type[Any]]] = [
    ("google_ads", GoogleAdsAdapter),
]


def _now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def _short_token() -> str:
    """Short random suffix for backup directory names."""
    import secrets

    return secrets.token_hex(4)


def _sha256_file(path: Path) -> str:
    import hashlib

    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


_MAX_BUNDLE_BYTES = 50 * 1024 * 1024  # 50 MB — generous cap, defense vs zip bombs


def _load_workbook(src: Path) -> Workbook:
    """Open the xlsx as a read-only openpyxl workbook.

    Read-only mode is faster and lower-memory; we only need values,
    not formulas / styles / images.

    Raises :class:`BundleImportError` for files larger than
    ``_MAX_BUNDLE_BYTES`` (defense in depth against zip-bomb XLSX
    files), and for files that openpyxl rejects as malformed. Other
    exceptions (MemoryError, OSError(EMFILE), etc.) are NOT caught
    here so the user sees the real cause.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError as exc:
        raise BundleImportError(
            "openpyxl is required for bundle import. "
            "Install with: pip install 'openpyxl>=3.1,<4'"
        ) from exc

    import zipfile

    size = src.stat().st_size
    if size > _MAX_BUNDLE_BYTES:
        raise BundleImportError(
            f"{src.name}: file is {size:,} bytes (limit "
            f"{_MAX_BUNDLE_BYTES:,}). The mureo Sheet template "
            "produces files well under this limit; an XLSX this "
            "large is likely not a mureo bundle."
        )

    try:
        return load_workbook(src, read_only=True, data_only=True)
    except (zipfile.BadZipFile, InvalidFileException, OSError) as exc:
        raise BundleImportError(
            f"{src.name}: failed to open as XLSX ({type(exc).__name__}: "
            f"{exc}). Verify the file is a valid Microsoft Excel "
            "workbook (File → Download → Microsoft Excel from your "
            "mureo Sheet)."
        ) from exc


def import_bundle(
    src: Path,
    *,
    replace: bool = False,
) -> dict[str, Any]:
    """Import a single XLSX bundle into ``~/.mureo/byod/``.

    Args:
        src: Path to the XLSX file exported from the mureo Sheet
            template.
        replace: When True, overwrite existing BYOD entries for any
            platform present in the bundle. When False (default),
            refuse if any target platform already has BYOD data.

    Returns:
        Mapping of platform → manifest entry for every platform that
        was successfully imported.

    Raises:
        BundleImportError: when the file is missing, is not a valid
            XLSX, contains no recognized tabs, conflicts with existing
            BYOD entries (without ``replace``), or when any
            dispatched adapter fails after partial files have been
            written (a rollback is attempted before re-raising).
    """
    from pathlib import Path as _Path

    src = _Path(src).expanduser().resolve()
    if not src.is_file():
        raise BundleImportError(f"{src}: file not found")

    workbook = _load_workbook(src)
    try:
        # Discover which platforms have data before touching disk.
        dispatched: list[tuple[str, type[Any]]] = []
        for platform, adapter_cls in _BUNDLE_ADAPTERS:
            if adapter_cls.has_tab(workbook):
                dispatched.append((platform, adapter_cls))

        if not dispatched:
            tabs = list(workbook.sheetnames)
            raise BundleImportError(
                f"{src.name}: no recognized tabs found. "
                f"Workbook has: {tabs}. "
                f"Expected a 'campaigns' tab (plus optionally "
                f"ad_groups / search_terms / keywords / "
                f"auction_insights) produced by the mureo "
                f"Google Ads Script."
            )

        manifest = read_manifest() or _empty_manifest()
        for platform, _ in dispatched:
            if platform in manifest["platforms"] and not replace:
                raise BundleImportError(
                    f"BYOD data for {platform!r} already exists. "
                    "Re-run with --replace to overwrite, or "
                    f"`mureo byod remove --{platform.replace('_', '-')}` "
                    "first."
                )

        src_sha = _sha256_file(src)
        byod_root = byod_data_dir().resolve()
        byod_root.mkdir(parents=True, exist_ok=True)

        # Per-platform backups so a later adapter's failure restores any
        # previously valid import for *earlier* platforms in the same run.
        # Without this, `--replace` over a 2-platform bundle where the
        # 2nd adapter raises would leave the user with neither the new
        # data nor their previous data for platform 1.
        prior_manifest = read_manifest()
        backups: dict[str, Path] = {}
        new_dirs: list[Path] = []
        results: dict[str, Any] = {}
        try:
            for platform, adapter_cls in dispatched:
                dst_dir = (byod_data_dir() / platform).resolve()
                # Defense-in-depth: refuse to write outside ~/.mureo/byod/.
                # platform values come from the closed _BUNDLE_ADAPTERS
                # constant today, so this guard is a no-op on the current
                # call path. It exists so a future contributor accepting
                # an external `platform` value cannot escape the BYOD
                # root via path traversal (../, absolute paths, etc.).
                if byod_root != dst_dir and byod_root not in dst_dir.parents:
                    raise BundleImportError(
                        f"Refusing to write outside BYOD root: {dst_dir}"
                    )
                if dst_dir.exists():
                    backup = dst_dir.with_suffix(f".backup-{_short_token()}")
                    dst_dir.rename(backup)
                    backups[platform] = backup
                new_dirs.append(dst_dir)

                try:
                    result = adapter_cls().normalize_from_workbook(workbook, dst_dir)
                except UnsupportedFormatError as exc:
                    # Wrap adapter validation errors so callers only need
                    # to handle BundleImportError. Rollback below cleans
                    # any partial CSVs from this and earlier adapters.
                    raise BundleImportError(f"{platform}: {exc}") from exc
                entry = {
                    "files": list(result.files_written),
                    "date_range": {
                        "start": result.date_range[0],
                        "end": result.date_range[1],
                    },
                    "rows": result.rows,
                    "campaigns": result.campaigns,
                    "ad_groups": result.ad_groups,
                    "source_format": result.source_format,
                    "imported_at": _now_iso(),
                    "source_file_sha256": src_sha,
                    "source_filename": src.name,
                }
                results[platform] = entry
                manifest["platforms"][platform] = entry
        except Exception:
            # Roll back: remove any new dirs, restore backups, restore
            # the prior manifest so disk + manifest stay consistent.
            # Each step is best-effort but logs on failure so the user
            # has a diagnostic trail if rollback itself partially fails
            # (e.g., a broken symlink or stale FD on the platform dir).
            for d in new_dirs:
                if d.exists():
                    try:
                        shutil.rmtree(d)
                    except OSError as roll_exc:
                        logger.warning(
                            "rollback: could not remove new dir %s: %s",
                            d,
                            roll_exc,
                        )
            for platform_name, backup in backups.items():
                target = byod_data_dir() / platform_name
                if backup.exists():
                    try:
                        backup.rename(target)
                    except OSError as roll_exc:
                        logger.warning(
                            "rollback: could not restore backup %s -> %s: %s. "
                            "Prior data is still present in the backup "
                            "directory; restore manually if needed.",
                            backup,
                            target,
                            roll_exc,
                        )
            if prior_manifest is not None:
                try:
                    write_manifest(prior_manifest)
                except OSError as roll_exc:
                    logger.warning(
                        "rollback: could not restore prior manifest: %s",
                        roll_exc,
                    )
            raise

        # Success: drop the backups (we have new data on disk).
        for backup in backups.values():
            if backup.exists():
                try:
                    shutil.rmtree(backup)
                except OSError as cleanup_exc:
                    logger.warning(
                        "could not remove backup %s after successful "
                        "import: %s. Safe to delete by hand.",
                        backup,
                        cleanup_exc,
                    )

        manifest["imported_on"] = _now_iso()
        write_manifest(manifest)
        return results
    finally:
        # Read-only workbooks hold a file handle until closed.
        with contextlib.suppress(Exception):
            workbook.close()


def _empty_manifest() -> dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "imported_on": _now_iso(),
        "platforms": {},
    }
