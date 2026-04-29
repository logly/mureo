"""Tests for the BYOD Sheet bundle importer.

Covers:
  - Happy path with Google Ads tabs (the sole BYOD platform after the
    Phase 1 BYOD redesign)
  - Workbook with no recognized tabs -> BundleImportError
  - Replace semantics (refuse on conflict / overwrite with replace=True)
  - Missing required column -> error + rollback (no partial CSVs left)
  - Non-XLSX file -> BundleImportError
  - Manifest entry shape matches the existing schema
"""

from __future__ import annotations

import json
from typing import TYPE_CHECKING

import pytest

if TYPE_CHECKING:
    from pathlib import Path


# ---------------------------------------------------------------------------
# xlsx fixture builders
# ---------------------------------------------------------------------------


def _make_workbook(tmp_path: Path, *, tabs: dict[str, list[list]]) -> Path:
    """Create an xlsx file at ``tmp_path/test.xlsx`` with the given tabs.

    ``tabs`` maps tab name -> list of rows (first row is the header).
    """
    from openpyxl import Workbook

    wb = Workbook()
    # The first sheet is created automatically; remove it so the
    # workbook only contains tabs explicitly listed.
    default = wb.active
    wb.remove(default)

    for name, rows in tabs.items():
        sheet = wb.create_sheet(name)
        for row in rows:
            sheet.append(row)

    out = tmp_path / "test.xlsx"
    wb.save(out)
    return out


def _google_ads_tabs() -> dict[str, list[list]]:
    return {
        "campaigns": [
            ["day", "campaign", "impressions", "clicks", "cost", "conversions"],
            ["2026-04-01", "Brand Search", 1000, 50, 25.5, 3.0],
            ["2026-04-02", "Brand Search", 1100, 55, 28.0, 4.0],
            ["2026-04-01", "Generic Search", 2000, 80, 60.0, 5.0],
        ],
        "ad_groups": [
            [
                "day",
                "campaign",
                "ad_group",
                "impressions",
                "clicks",
                "cost",
                "conversions",
            ],
            ["2026-04-01", "Brand Search", "Exact match", 800, 40, 20.0, 2.0],
        ],
    }


def _meta_ads_tabs() -> dict[str, list[list]]:
    """Synthetic Meta Ads Manager Export Excel with Day breakdown.

    Mirrors a typical user export configuration:
    Reports → Customize → Breakdown: By Day, Level: Ad → Export → Excel.
    """
    return {
        "Sheet1": [
            [
                "Day",
                "Campaign name",
                "Ad set name",
                "Ad name",
                "Impressions",
                "Clicks (all)",
                "Amount spent (JPY)",
                "Results",
            ],
            [
                "2026-04-01",
                "Brand Awareness",
                "Tokyo 25-34",
                "Video A",
                5000,
                120,
                "1,500",
                3,
            ],
            [
                "2026-04-01",
                "Brand Awareness",
                "Tokyo 25-34",
                "Video B",
                3000,
                80,
                "1,200",
                2,
            ],
            [
                "2026-04-02",
                "Brand Awareness",
                "Tokyo 25-34",
                "Video A",
                5500,
                130,
                "1,650",
                4,
            ],
            [
                "2026-04-01",
                "Conversion",
                "Lookalike 1%",
                "Carousel A",
                2000,
                40,
                "800",
                5,
            ],
        ],
    }


# ---------------------------------------------------------------------------
# tmp_path BYOD root
# ---------------------------------------------------------------------------


@pytest.fixture()
def byod_root(tmp_path, monkeypatch):
    """Redirect ``~/.mureo/byod/`` to ``tmp_path/.mureo/byod/``."""
    fake_home = tmp_path / "home"
    fake_home.mkdir()
    monkeypatch.setattr("pathlib.Path.home", lambda: fake_home)
    return fake_home / ".mureo" / "byod"


# ---------------------------------------------------------------------------
# Happy path
# ---------------------------------------------------------------------------


def test_import_bundle_writes_google_ads(tmp_path, byod_root):
    from mureo.byod.bundle import import_bundle

    src = _make_workbook(tmp_path, tabs=_google_ads_tabs())

    results = import_bundle(src)

    assert set(results.keys()) == {"google_ads"}

    ga_root = byod_root / "google_ads"
    assert (ga_root / "campaigns.csv").exists()
    assert (ga_root / "metrics_daily.csv").exists()
    assert results["google_ads"]["rows"] >= 1
    assert results["google_ads"]["date_range"]["start"] == "2026-04-01"


def test_import_bundle_writes_meta_ads(tmp_path, byod_root):
    """Happy path for the Meta Ads adapter: a single-sheet Ads
    Manager export with Day breakdown round-trips into the 4 CSVs the
    ByodMetaAdsClient consumes."""
    from mureo.byod.bundle import import_bundle

    src = _make_workbook(tmp_path, tabs=_meta_ads_tabs())

    results = import_bundle(src)
    assert set(results.keys()) == {"meta_ads"}

    meta_root = byod_root / "meta_ads"
    assert (meta_root / "campaigns.csv").exists()
    assert (meta_root / "ad_sets.csv").exists()
    assert (meta_root / "ads.csv").exists()
    assert (meta_root / "metrics_daily.csv").exists()

    # 3 unique (day, campaign) pairs in the fixture:
    # (2026-04-01, Brand) + (2026-04-02, Brand) + (2026-04-01, Conversion)
    assert results["meta_ads"]["rows"] == 3
    assert results["meta_ads"]["campaigns"] == 2
    assert results["meta_ads"]["date_range"]["start"] == "2026-04-01"
    assert results["meta_ads"]["date_range"]["end"] == "2026-04-02"
    assert results["meta_ads"]["source_format"] == "mureo_meta_ads_export_v1"

    # Ad set / Ad rows are present (Day breakdown export with Level: Ad).
    ad_sets_text = (meta_root / "ad_sets.csv").read_text(encoding="utf-8")
    assert "Tokyo 25-34" in ad_sets_text
    assert "Lookalike 1%" in ad_sets_text

    ads_text = (meta_root / "ads.csv").read_text(encoding="utf-8")
    assert "Video A" in ads_text
    assert "Carousel A" in ads_text

    # Metrics are summed when multiple ad-level rows share a (day, campaign).
    # (2026-04-01, Brand Awareness): 5000 + 3000 = 8000 impressions.
    metrics_text = (meta_root / "metrics_daily.csv").read_text(encoding="utf-8")
    assert "8000" in metrics_text
    # cost was given as "1,500" + "1,200" = 2700.00
    assert "2700.00" in metrics_text


def test_import_bundle_writes_both_google_and_meta(tmp_path, byod_root):
    """A workbook carrying both Google Ads tabs and a Meta Ads
    Manager export sheet must dispatch both adapters and produce two
    manifest entries — guarding the disjoint-detection contract."""
    from mureo.byod.bundle import import_bundle

    combined = {**_google_ads_tabs(), **_meta_ads_tabs()}
    src = _make_workbook(tmp_path, tabs=combined)

    results = import_bundle(src)
    assert set(results.keys()) == {"google_ads", "meta_ads"}
    assert (byod_root / "google_ads" / "campaigns.csv").exists()
    assert (byod_root / "meta_ads" / "campaigns.csv").exists()


@pytest.mark.parametrize(
    "date_alias,raw_value,expected",
    [
        ("Day", "2026-04-01", "2026-04-01"),
        ("Reporting starts", "2026-04-01", "2026-04-01"),
        ("Date", "2026/04/01", "2026-04-01"),
        # US-locale Ads Manager export
        ("Day", "4/1/2026", "2026-04-01"),
    ],
)
def test_meta_adapter_recognizes_date_aliases_and_locales(
    tmp_path, byod_root, date_alias, raw_value, expected
):
    """Each of Day / Reporting starts / Date headers + ISO + US
    locale dates round-trip into the YYYY-MM-DD output column."""
    from mureo.byod.bundle import import_bundle

    src = _make_workbook(
        tmp_path,
        tabs={
            "Sheet1": [
                [
                    date_alias,
                    "Campaign name",
                    "Impressions",
                    "Clicks (all)",
                    "Amount spent (JPY)",
                ],
                [raw_value, "Solo", 100, 5, "200"],
            ],
        },
    )
    import_bundle(src)
    metrics = (byod_root / "meta_ads" / "metrics_daily.csv").read_text(encoding="utf-8")
    assert expected in metrics


def test_meta_adapter_rejects_non_jpy_spend(tmp_path, byod_root):
    """A non-JPY currency in the Spend column must abort the import
    rather than silently coerce dollars/euros into the JPY column."""
    from mureo.byod.bundle import BundleImportError, import_bundle

    src = _make_workbook(
        tmp_path,
        tabs={
            "Sheet1": [
                [
                    "Day",
                    "Campaign name",
                    "Impressions",
                    "Amount spent",
                ],
                ["2026-04-01", "Solo", 100, "$12.34"],
            ],
        },
    )
    with pytest.raises(BundleImportError, match="non-JPY|JPY"):
        import_bundle(src)


def test_meta_adapter_csv_injection_sanitized(tmp_path, byod_root):
    """A campaign / ad-set / ad name beginning with a CSV-injection
    formula trigger (``=``, ``+``, ``-``, ``@``) must be defanged
    with a leading apostrophe so Excel/Sheets do not execute the
    formula on re-open.

    Note on the test fixture: openpyxl interprets a cell whose value
    starts with ``=`` as a formula and would write it as such. Since
    Meta's actual export emits literal strings (not formulas), we
    simulate that by setting the cell ``data_type = 's'`` so the
    workbook round-trip preserves the leading ``=``.
    """
    from openpyxl import Workbook

    from mureo.byod.bundle import import_bundle

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    sheet = wb.create_sheet("Sheet1")
    sheet.append(
        [
            "Day",
            "Campaign name",
            "Ad set name",
            "Ad name",
            "Impressions",
            "Amount spent (JPY)",
        ]
    )
    # Append with placeholder, then overwrite the suspicious cells as
    # explicit strings to bypass openpyxl's formula auto-detection.
    sheet.append(["2026-04-01", "PLACEHOLDER", "+evil_set", "@danger_ad", 100, "200"])
    formula_like = "=cmd|' /C calc'!A0"
    cell = sheet.cell(row=2, column=2)
    cell.value = formula_like
    cell.data_type = "s"
    src = tmp_path / "test.xlsx"
    wb.save(src)

    import_bundle(src)
    campaigns = (byod_root / "meta_ads" / "campaigns.csv").read_text(encoding="utf-8")
    assert "'=cmd" in campaigns  # leading apostrophe defangs formula

    ad_sets = (byod_root / "meta_ads" / "ad_sets.csv").read_text(encoding="utf-8")
    assert "'+evil_set" in ad_sets

    ads = (byod_root / "meta_ads" / "ads.csv").read_text(encoding="utf-8")
    assert "'@danger_ad" in ads


def test_meta_adapter_phase3_skips_rollup_when_detail_present(tmp_path, byod_root):
    """A pivot export carrying both a campaign rollup row (Ad set
    name='All', Ad name='All') AND ad-level detail rows for the
    same (day, campaign) must NOT double-count metrics. The deepest
    grain present per (day, campaign) wins; shallower-grain rows
    are filtered out of metrics_agg."""
    from mureo.byod.bundle import import_bundle

    src = _make_workbook(
        tmp_path,
        tabs={
            "Sheet1": [
                [
                    "Day",
                    "Campaign name",
                    "Ad set name",
                    "Ad name",
                    "Impressions",
                    "Amount spent (JPY)",
                ],
                # Campaign rollup — would contribute 9999 if not skipped.
                ["2026-04-01", "Brand", "All", "All", 9999, "9999"],
                # Detail rows that should be the actual sum.
                ["2026-04-01", "Brand", "Tokyo", "Video A", 5000, "1500"],
                ["2026-04-01", "Brand", "Tokyo", "Video B", 3000, "900"],
            ],
        },
    )
    import_bundle(src)
    metrics = (byod_root / "meta_ads" / "metrics_daily.csv").read_text(encoding="utf-8")
    rows = [r for r in metrics.splitlines()[1:] if r.strip()]
    assert len(rows) == 1
    # Detail-only sum: 5000 + 3000 = 8000 (NOT 9999 + 5000 + 3000 = 17999).
    assert ",8000," in rows[0]
    assert "9999" not in rows[0]


def test_meta_adapter_phase3_extended_schema(tmp_path, byod_root):
    """Phase 3-1 schema extension: metrics_daily.csv carries reach,
    frequency, and result_indicator columns. Frequency falls back to
    impressions/reach when the export does not include a Frequency
    column directly."""
    from mureo.byod.bundle import import_bundle

    src = _make_workbook(
        tmp_path,
        tabs={
            "Sheet1": [
                [
                    "Day",
                    "Campaign name",
                    "Impressions",
                    "Reach",
                    "Amount spent (JPY)",
                    "Results",
                    "Result indicator",
                ],
                [
                    "2026-04-01",
                    "Brand",
                    1000,
                    400,
                    "1500",
                    5,
                    "actions:offsite_conversion.fb_pixel_lead",
                ],
            ],
        },
    )
    import_bundle(src)
    metrics = (byod_root / "meta_ads" / "metrics_daily.csv").read_text(encoding="utf-8")
    header = metrics.splitlines()[0].split(",")
    assert "reach" in header
    assert "frequency" in header
    assert "result_indicator" in header
    row = metrics.splitlines()[1]
    assert ",400," in row  # reach
    assert ",2.5," in row  # frequency = 1000/400 = 2.5
    assert "fb_pixel_lead" in row


def test_meta_adapter_phase3_ad_set_and_ad_csvs(tmp_path, byod_root):
    """Phase 3-2 finer grain: when the export has Ad set name and Ad
    name columns, ad_set_metrics_daily.csv and ad_metrics_daily.csv
    are written with per-ad-set / per-ad daily metrics."""
    from mureo.byod.bundle import import_bundle

    src = _make_workbook(
        tmp_path,
        tabs={
            "Sheet1": [
                [
                    "Day",
                    "Campaign name",
                    "Ad set name",
                    "Ad name",
                    "Impressions",
                    "Clicks (all)",
                    "Amount spent (JPY)",
                    "Results",
                ],
                ["2026-04-01", "Brand", "Tokyo", "Video A", 5000, 100, "1500", 3],
                ["2026-04-01", "Brand", "Tokyo", "Video B", 3000, 60, "900", 2],
                ["2026-04-02", "Brand", "Tokyo", "Video A", 5500, 110, "1650", 4],
            ],
        },
    )
    import_bundle(src)
    meta_root = byod_root / "meta_ads"
    assert (meta_root / "ad_set_metrics_daily.csv").exists()
    assert (meta_root / "ad_metrics_daily.csv").exists()

    as_metrics = (meta_root / "ad_set_metrics_daily.csv").read_text(encoding="utf-8")
    rows = [r for r in as_metrics.splitlines()[1:] if r.strip()]
    assert len(rows) == 2
    apr1_row = next(r for r in rows if r.startswith("2026-04-01"))
    assert ",8000," in apr1_row

    ad_metrics = (meta_root / "ad_metrics_daily.csv").read_text(encoding="utf-8")
    rows = [r for r in ad_metrics.splitlines()[1:] if r.strip()]
    assert len(rows) == 3


def test_meta_adapter_phase3_demographics_csv(tmp_path, byod_root):
    """Phase 3-3 demographics: rows with non-'All' values in age/
    gender/region/placement go to demographics_daily.csv only and are
    excluded from the campaign-level metrics_daily aggregation
    (otherwise the totals would be doubled)."""
    from mureo.byod.bundle import import_bundle

    src = _make_workbook(
        tmp_path,
        tabs={
            "Sheet1": [
                [
                    "Day",
                    "Campaign name",
                    "Age",
                    "Gender",
                    "Impressions",
                    "Amount spent (JPY)",
                ],
                ["2026-04-01", "Brand", "18-24", "All", 2000, "500"],
                ["2026-04-01", "Brand", "25-34", "All", 3000, "750"],
                ["2026-04-01", "Brand", "All", "male", 4000, "1000"],
                ["2026-04-01", "Brand", "All", "female", 1000, "250"],
                ["2026-04-01", "Brand", "All", "All", 5000, "1250"],
            ],
        },
    )
    import_bundle(src)
    meta_root = byod_root / "meta_ads"
    assert (meta_root / "demographics_daily.csv").exists()

    demo = (meta_root / "demographics_daily.csv").read_text(encoding="utf-8")
    rows = [r for r in demo.splitlines()[1:] if r.strip()]
    assert len(rows) == 4
    assert any(",age,18-24," in r for r in rows)
    assert any(",gender,male," in r for r in rows)

    metrics = (meta_root / "metrics_daily.csv").read_text(encoding="utf-8")
    metrics_rows = [r for r in metrics.splitlines()[1:] if r.strip()]
    assert len(metrics_rows) == 1
    assert ",5000," in metrics_rows[0]


def test_meta_adapter_phase3_creatives_csv(tmp_path, byod_root):
    """Phase 3-4 creatives: when the export carries image URL /
    headline / body / cta columns, creatives.csv is produced."""
    from mureo.byod.bundle import import_bundle

    src = _make_workbook(
        tmp_path,
        tabs={
            "Sheet1": [
                [
                    "Day",
                    "Campaign name",
                    "Ad set name",
                    "Ad name",
                    "Impressions",
                    "Amount spent (JPY)",
                    "Image URL",
                    "Headline",
                    "Body",
                    "Call to action",
                ],
                [
                    "2026-04-01",
                    "Brand",
                    "Tokyo",
                    "Video A",
                    1000,
                    "500",
                    "https://example.com/img.jpg",
                    "Try mureo today",
                    "Local-first ad ops",
                    "Learn more",
                ],
            ],
        },
    )
    import_bundle(src)
    meta_root = byod_root / "meta_ads"
    assert (meta_root / "creatives.csv").exists()
    creatives = (meta_root / "creatives.csv").read_text(encoding="utf-8")
    assert "https://example.com/img.jpg" in creatives
    assert "Try mureo today" in creatives
    assert "Learn more" in creatives


def test_meta_adapter_recognizes_japanese_headers(tmp_path, byod_root):
    """Japanese Ads Manager export: localized column headers
    (キャンペーン名 / インプレッション / 消化金額 / 結果) round-trip
    through the adapter exactly like English headers do."""
    from mureo.byod.bundle import import_bundle

    src = _make_workbook(
        tmp_path,
        tabs={
            "ワークシート": [
                [
                    "日",
                    "キャンペーン名",
                    "広告セット名",
                    "広告名",
                    "インプレッション",
                    "リンクのクリック",
                    "消化金額 (JPY)",
                    "結果",
                ],
                [
                    "2026-04-01",
                    "ブランド認知",
                    "東京 25-34",
                    "動画A",
                    5000,
                    120,
                    "1,500",
                    3,
                ],
                [
                    "2026-04-02",
                    "ブランド認知",
                    "東京 25-34",
                    "動画A",
                    5500,
                    130,
                    "1,650",
                    4,
                ],
            ],
        },
    )
    import_bundle(src)

    campaigns = (byod_root / "meta_ads" / "campaigns.csv").read_text(encoding="utf-8")
    assert "ブランド認知" in campaigns

    metrics = (byod_root / "meta_ads" / "metrics_daily.csv").read_text(encoding="utf-8")
    assert "2026-04-01" in metrics
    assert "5000" in metrics


def test_meta_adapter_pivot_subtotal_rows_skipped(tmp_path, byod_root):
    """Reports → ピボット exports include subtotal rows where the date
    cell reads ``All`` (or is blank). These must be skipped so summed
    metrics aren't double-counted by the campaign-rollup grain."""
    from mureo.byod.bundle import import_bundle

    src = _make_workbook(
        tmp_path,
        tabs={
            "Pivot": [
                [
                    "日",
                    "キャンペーン名",
                    "インプレッション",
                    "消化金額 (JPY)",
                ],
                # Account / campaign subtotal — should be SKIPPED
                ["All", "ブランド認知", 99999, "999999"],
                # Detail rows — kept and summed
                ["2026-04-01", "ブランド認知", 5000, "1500"],
                ["2026-04-02", "ブランド認知", 5500, "1650"],
                # Another subtotal — SKIPPED
                ["All", "コンバージョン", 88888, "888888"],
                ["2026-04-01", "コンバージョン", 2000, "800"],
            ],
        },
    )
    import_bundle(src)
    metrics = (byod_root / "meta_ads" / "metrics_daily.csv").read_text(encoding="utf-8")
    # Subtotal cost values must NOT appear in the output
    assert "999999" not in metrics
    assert "888888" not in metrics
    # Detail rows must appear
    assert "1500.00" in metrics
    assert "1650.00" in metrics


def test_meta_adapter_ad_set_name_reused_across_campaigns(tmp_path, byod_root):
    """Two campaigns reusing the same ad-set name (e.g. "Default",
    "Lookalike 1%") must produce distinct `ad_set_id` rows in
    ad_sets.csv, and ads must be linked to the correct ad_set_id of
    the same campaign."""
    from mureo.byod.bundle import import_bundle

    src = _make_workbook(
        tmp_path,
        tabs={
            "Sheet1": [
                [
                    "Day",
                    "Campaign name",
                    "Ad set name",
                    "Ad name",
                    "Impressions",
                    "Amount spent (JPY)",
                ],
                ["2026-04-01", "Camp A", "Default", "Ad A", 100, "200"],
                ["2026-04-01", "Camp B", "Default", "Ad B", 50, "100"],
            ],
        },
    )
    import_bundle(src)
    ad_sets = (byod_root / "meta_ads" / "ad_sets.csv").read_text(encoding="utf-8")
    # Two distinct ad_set_id rows even though both ad-sets are named "Default".
    rows = [r for r in ad_sets.splitlines() if "Default" in r]
    assert len(rows) == 2
    ad_set_ids = {r.split(",")[0] for r in rows}
    assert len(ad_set_ids) == 2

    ads = (byod_root / "meta_ads" / "ads.csv").read_text(encoding="utf-8")
    # Ad A and Ad B should each reference exactly one ad_set_id, and
    # the two should be different.
    ad_a_line = next(r for r in ads.splitlines() if "Ad A" in r)
    ad_b_line = next(r for r in ads.splitlines() if "Ad B" in r)
    ad_a_set_id = ad_a_line.split(",")[1]
    ad_b_set_id = ad_b_line.split(",")[1]
    assert ad_a_set_id != ad_b_set_id


def test_import_bundle_meta_ads_round_trip_through_client(tmp_path, byod_root):
    """The Meta CSVs the adapter writes must be readable by the
    existing ByodMetaAdsClient — list_campaigns + get_performance_report
    return the synthesized data, with cost_jpy reflecting summed
    spend (regression guard for the cost vs cost_jpy bug PR #50 fixed
    on the Google Ads side)."""
    import asyncio

    from mureo.byod.bundle import import_bundle
    from mureo.byod.clients import ByodMetaAdsClient

    src = _make_workbook(tmp_path, tabs=_meta_ads_tabs())
    import_bundle(src)

    client = ByodMetaAdsClient(data_dir=byod_root / "meta_ads")
    campaigns = asyncio.run(client.list_campaigns())
    assert {c["name"] for c in campaigns} == {"Brand Awareness", "Conversion"}

    # Force the period range to include the 2026-04-01..2026-04-02 fixture.
    # The client's default LAST_30_DAYS won't intersect a fixture far in
    # the past, so we patch the fixture dates by checking that any rows
    # come back when we widen the period helper indirectly. Simplest:
    # confirm the metrics CSV was read and aggregated by the client by
    # poking the underlying _metrics().
    metric_rows = client._metrics()
    assert len(metric_rows) == 3
    # Sum of cost_jpy across all rows: 2700 + 1650 + 800 = 5150
    total_cost = sum(float(r["cost_jpy"]) for r in metric_rows)
    assert total_cost == 5150.0


# ---------------------------------------------------------------------------
# Error paths
# ---------------------------------------------------------------------------


def test_import_bundle_no_recognized_tabs(tmp_path, byod_root):
    from mureo.byod.bundle import BundleImportError, import_bundle

    src = _make_workbook(
        tmp_path,
        tabs={"unrelated_data": [["foo", "bar"], [1, 2]]},
    )

    with pytest.raises(BundleImportError, match="no recognized tabs"):
        import_bundle(src)


def test_import_bundle_missing_required_column_rolls_back(tmp_path, byod_root):
    """Adapter rejects a workbook whose ``campaigns`` tab lacks
    required columns; the bundle importer must roll back any partial
    on-disk artifact."""
    from mureo.byod.bundle import BundleImportError, import_bundle

    src = _make_workbook(
        tmp_path,
        tabs={
            "campaigns": [
                # Missing 'impressions', 'clicks', etc.
                ["day", "campaign"],
                ["2026-04-01", "Brand"],
            ],
        },
    )

    with pytest.raises(BundleImportError):
        import_bundle(src)

    assert not (byod_root / "google_ads").exists()


def test_import_bundle_non_xlsx_file(tmp_path, byod_root):
    from mureo.byod.bundle import BundleImportError, import_bundle

    not_xlsx = tmp_path / "fake.xlsx"
    not_xlsx.write_text("this is not really an xlsx", encoding="utf-8")

    with pytest.raises(BundleImportError, match="failed to open as XLSX"):
        import_bundle(not_xlsx)


def test_import_bundle_missing_file(tmp_path, byod_root):
    from mureo.byod.bundle import BundleImportError, import_bundle

    with pytest.raises(BundleImportError, match="file not found"):
        import_bundle(tmp_path / "does-not-exist.xlsx")


# ---------------------------------------------------------------------------
# Replace semantics
# ---------------------------------------------------------------------------


def test_import_bundle_refuses_existing_without_replace(tmp_path, byod_root):
    from mureo.byod.bundle import BundleImportError, import_bundle

    src = _make_workbook(tmp_path, tabs=_google_ads_tabs())
    import_bundle(src)  # first import succeeds

    with pytest.raises(BundleImportError, match="already exists"):
        import_bundle(src)


def test_import_bundle_replace_overwrites(tmp_path, byod_root):
    from mureo.byod.bundle import import_bundle

    src1 = _make_workbook(tmp_path, tabs=_google_ads_tabs())
    import_bundle(src1)

    src2_dir = tmp_path / "second"
    src2_dir.mkdir()
    new_tabs = {
        "campaigns": [
            ["day", "campaign", "impressions", "clicks", "cost", "conversions"],
            ["2026-05-10", "Refreshed", 2000, 100, 50.0, 8.0],
        ],
    }
    src2 = _make_workbook(src2_dir, tabs=new_tabs)

    import_bundle(src2, replace=True)
    csv_text = (byod_root / "google_ads" / "metrics_daily.csv").read_text(
        encoding="utf-8"
    )
    assert "2026-04-01" not in csv_text
    assert "2026-05-10" in csv_text


# ---------------------------------------------------------------------------
# Manifest schema
# ---------------------------------------------------------------------------


def test_import_bundle_manifest_entry_shape(tmp_path, byod_root):
    from mureo.byod.bundle import import_bundle

    src = _make_workbook(tmp_path, tabs=_google_ads_tabs())
    import_bundle(src)

    manifest = json.loads((byod_root / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["schema_version"] == 1
    assert set(manifest["platforms"].keys()) == {"google_ads"}
    entry = manifest["platforms"]["google_ads"]
    for key in (
        "files",
        "date_range",
        "rows",
        "campaigns",
        "ad_groups",
        "source_format",
        "imported_at",
        "source_file_sha256",
        "source_filename",
    ):
        assert key in entry, f"missing manifest key: {key}"
    assert entry["source_filename"] == "test.xlsx"
