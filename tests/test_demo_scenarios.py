"""Contract tests for the demo scenario registry.

Each registered scenario must:
  - parse cleanly through ``mureo.context.state.parse_state``
  - have its bundle round-trip through ``mureo.byod.bundle.import_bundle``
  - have STATE.json campaign_ids that equal the BYOD adapter's
    synthesized ids (so workflow skills can join the two)

These tests run once per registered scenario, so adding a new scenario
automatically gets the same coverage without copy/pasting test bodies.
"""

from __future__ import annotations

import json
from typing import TYPE_CHECKING

import pytest

if TYPE_CHECKING:
    from pathlib import Path


pytestmark = pytest.mark.unit


@pytest.fixture(autouse=True)
def byod_root(tmp_path, monkeypatch):
    """Sandbox ``~/.mureo`` for every test in this module."""
    fake_home = tmp_path / "home"
    fake_home.mkdir()
    monkeypatch.setattr("pathlib.Path.home", lambda: fake_home)
    return fake_home


# ---------------------------------------------------------------------------
# Registry
# ---------------------------------------------------------------------------


def test_registry_has_default_scenario() -> None:
    from mureo.demo.scenarios import DEFAULT_SCENARIO, SCENARIOS

    assert DEFAULT_SCENARIO in SCENARIOS


def test_registry_default_is_seasonality_trap() -> None:
    """The chosen Stage-1 default is the FlavorBox Seasonality Trap.

    This pin is intentional — it's the story we're optimizing the demo
    around. Changing the default should be a deliberate decision, not
    a refactor accident.
    """
    from mureo.demo.scenarios import DEFAULT_SCENARIO

    assert DEFAULT_SCENARIO == "seasonality-trap"


def test_get_scenario_resolves_default_when_none() -> None:
    from mureo.demo.scenarios import DEFAULT_SCENARIO, get_scenario

    sc = get_scenario(None)
    assert sc.name == DEFAULT_SCENARIO


def test_get_scenario_unknown_raises_clear_error() -> None:
    from mureo.demo.scenarios import get_scenario

    with pytest.raises(ValueError) as excinfo:
        get_scenario("does-not-exist")
    msg = str(excinfo.value)
    assert "does-not-exist" in msg
    # The error must list valid scenarios so the user can recover.
    assert "seasonality-trap" in msg


# ---------------------------------------------------------------------------
# Scenario shape — checked once per registered scenario
# ---------------------------------------------------------------------------


def _all_scenarios():
    from mureo.demo.scenarios import SCENARIOS

    return list(SCENARIOS.values())


def _ids_for_scenarios(scs):
    return [s.name for s in scs]


@pytest.fixture(params=_all_scenarios(), ids=_ids_for_scenarios(_all_scenarios()))
def scenario(request):
    return request.param


def test_scenario_metadata_is_populated(scenario) -> None:
    assert scenario.name
    assert scenario.title
    assert scenario.blurb
    assert scenario.days > 0
    assert scenario.brand


def test_scenario_sheet_rows_have_expected_tabs(scenario) -> None:
    """Every scenario must populate the 5 BYOD-recognized tabs.

    The BYOD adapter dispatches on tab name, so dropping any of these
    quietly disables that platform's BYOD path.
    """
    expected = {"campaigns", "ad_groups", "search_terms", "keywords", "meta_ads"}
    assert set(scenario.sheet_rows.keys()) >= expected
    for tab, rows in scenario.sheet_rows.items():
        assert len(rows) >= 2, f"{tab}: needs header + at least one row"


def test_scenario_state_doc_parses_via_context_state(scenario) -> None:
    """STATE.json must round-trip through ``mureo.context.state.parse_state``.

    A scenario whose state_doc fails to parse would silently break
    every workflow skill that reads STATE.json — caught here instead
    of at runtime in Claude Code.
    """
    from mureo.context.state import parse_state

    text = json.dumps(scenario.state_doc)
    doc = parse_state(text)
    assert doc.version == "2"
    assert doc.platforms is not None
    assert "google_ads" in doc.platforms
    assert "meta_ads" in doc.platforms


def test_scenario_state_campaign_names_appear_in_bundle(scenario) -> None:
    """Every STATE.json campaign name must have rows in the bundle data.

    A scenario author can typo a name in STATE.json that won't match
    any row in the campaigns / meta_ads tabs. With per-tab synthesized
    campaign_ids, the resulting STATE entry would have no joinable
    performance data — the workflow skill silently sees an empty
    campaign and reports nothing. Fail loudly here instead.
    """
    state = scenario.state_doc

    gads_rows = scenario.sheet_rows["campaigns"]
    # campaigns tab schema: day, campaign, impressions, clicks, cost, conversions
    gads_names = {row[1] for row in gads_rows[1:]}

    meta_rows = scenario.sheet_rows["meta_ads"]
    # meta_ads tab schema: Day, Campaign name, Ad set name, Ad name, ...
    meta_names = {row[1] for row in meta_rows[1:]}

    for c in state["platforms"]["google_ads"]["campaigns"]:
        assert c["campaign_name"] in gads_names, (
            f"{scenario.name}: STATE google_ads campaign "
            f"{c['campaign_name']!r} has no rows in campaigns tab"
        )
    for c in state["platforms"]["meta_ads"]["campaigns"]:
        assert c["campaign_name"] in meta_names, (
            f"{scenario.name}: STATE meta_ads campaign "
            f"{c['campaign_name']!r} has no rows in meta_ads tab"
        )


def test_scenario_has_minimum_data_breadth(scenario) -> None:
    """Each scenario must carry enough data for every workflow skill.

    Without these floors, ``/search-term-cleanup`` could legitimately
    say "nothing to clean up", ``/budget-rebalance`` could see one
    campaign and decline to act, etc. The numbers are picked so that
    every skill in mureo's catalog sees enough data to reach a
    non-trivial conclusion.
    """
    rows = scenario.sheet_rows

    # Header is always row 0; subtract it.
    n_campaigns_unique = len({row[1] for row in rows["campaigns"][1:]})
    n_ad_groups_unique = len({(row[1], row[2]) for row in rows["ad_groups"][1:]})
    n_search_terms = len(rows["search_terms"]) - 1
    n_keywords = len(rows["keywords"]) - 1
    n_meta_ads_unique = len({row[3] for row in rows["meta_ads"][1:]})

    assert n_campaigns_unique >= 4, (
        f"{scenario.name}: needs >=4 Google campaigns "
        f"for /budget-rebalance to have signal (have {n_campaigns_unique})"
    )
    assert (
        n_ad_groups_unique >= 6
    ), f"{scenario.name}: needs >=6 ad groups (have {n_ad_groups_unique})"
    assert n_search_terms >= 15, (
        f"{scenario.name}: needs >=15 search terms for "
        f"/search-term-cleanup outlier detection (have {n_search_terms})"
    )
    assert n_keywords >= 8, f"{scenario.name}: needs >=8 keywords (have {n_keywords})"
    assert n_meta_ads_unique >= 4, (
        f"{scenario.name}: needs >=4 unique Meta ads for "
        f"/creative-refresh to have signal (have {n_meta_ads_unique})"
    )

    n_action_log = len(scenario.state_doc.get("action_log", []))
    # Scenarios that opt out via ``requires_action_log=False`` do so
    # because a sparse / absent action_log is itself part of their
    # diagnostic signal (see ``Scenario.requires_action_log`` docs).
    if scenario.requires_action_log:
        assert n_action_log >= 3, (
            f"{scenario.name}: needs >=3 action_log entries so "
            f"/learn and /sync-state have history to evaluate "
            f"(have {n_action_log})"
        )


def test_scenario_strategy_md_has_required_sections(scenario) -> None:
    """STRATEGY.md must include the load-bearing sections workflow skills read.

    ``/goal-review`` reads numeric goals; ``/competitive-scan`` and
    ``/daily-check`` read constraints; multiple commands branch on
    Operation Mode. A scenario missing any of these silently
    underperforms in the matching skill.
    """
    text = scenario.strategy_md.lower()
    assert (
        "goal" in text or "目標" in text
    ), f"{scenario.name}: STRATEGY.md missing Goals section"
    assert (
        "constraint" in text or "制約" in text
    ), f"{scenario.name}: STRATEGY.md missing Constraints section"
    assert (
        "operation mode" in text or "運用モード" in text
    ), f"{scenario.name}: STRATEGY.md missing Operation Mode"


def test_byod_synthetic_id_helpers_consistent() -> None:
    """The Google and Meta adapters' ``_synthetic_id`` must agree.

    Both BYOD adapters synthesize campaign_id from the same SHA-256
    formula (``mureo/byod/adapters/google_ads.py:100``,
    ``mureo/byod/adapters/meta_ads.py:475``). The demo's STATE.json
    uses the Google adapter's helper for both platforms — if the two
    formulas ever diverge, demo Meta IDs would silently mismatch the
    BYOD-imported Meta CSV IDs. Pin the contract here so divergence
    fails with an obvious message rather than a deep-stack assertion.
    """
    from mureo.byod.adapters.google_ads import (
        _synthetic_id as _gads_synthetic_id,
    )
    from mureo.byod.adapters.meta_ads import (
        _synthetic_id as _meta_synthetic_id,
    )

    samples = ("Brand - Exact", "FlavorBox", "敏感肌 Awareness", "")
    for name in samples:
        assert _gads_synthetic_id("camp", name) == _meta_synthetic_id(
            "camp", name
        ), f"_synthetic_id divergence for {name!r}"


def test_scenario_strategy_md_is_non_empty_markdown(scenario) -> None:
    text = scenario.strategy_md
    assert text.strip()
    assert "#" in text


def test_scenario_bundle_round_trips_via_byod_pipeline(
    scenario, tmp_path: Path, byod_root
) -> None:
    """Each scenario's sheet_rows must produce a workbook that imports
    cleanly through the BYOD pipeline used by every real user.
    """
    from openpyxl import Workbook

    from mureo.byod.bundle import import_bundle

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    for tab, rows in scenario.sheet_rows.items():
        sheet = wb.create_sheet(tab)
        for row in rows:
            sheet.append(list(row))
    out = tmp_path / f"{scenario.name}.xlsx"
    wb.save(out)

    results = import_bundle(out)
    assert "google_ads" in results
    assert "meta_ads" in results
    assert results["google_ads"]["rows"] > 0
    assert results["meta_ads"]["rows"] > 0


def test_scenario_state_campaign_ids_match_byod_csv(
    scenario, tmp_path: Path, byod_root
) -> None:
    """STATE.json campaign_ids must equal the BYOD adapter's synthesized ids.

    Workflow skills join STATE.json metadata against BYOD performance
    data on ``campaign_id``. Drift would silently break joins.
    """
    import csv as _csv
    from pathlib import Path as _Path

    from openpyxl import Workbook

    from mureo.byod.bundle import import_bundle

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    for tab, rows in scenario.sheet_rows.items():
        sheet = wb.create_sheet(tab)
        for row in rows:
            sheet.append(list(row))
    out = tmp_path / f"{scenario.name}.xlsx"
    wb.save(out)
    import_bundle(out)

    state = scenario.state_doc
    byod_root_dir = _Path.home() / ".mureo" / "byod"

    for platform in ("google_ads", "meta_ads"):
        state_ids = {
            c["campaign_name"]: c["campaign_id"]
            for c in state["platforms"][platform]["campaigns"]
        }
        csv_path = byod_root_dir / platform / "campaigns.csv"
        with csv_path.open(encoding="utf-8") as f:
            csv_ids = {row["name"]: row["campaign_id"] for row in _csv.DictReader(f)}
        for name, sid in state_ids.items():
            assert csv_ids.get(name) == sid, (
                f"{scenario.name}/{platform}: campaign_id mismatch for "
                f"{name!r}: STATE.json={sid!r} vs BYOD csv={csv_ids.get(name)!r}"
            )
